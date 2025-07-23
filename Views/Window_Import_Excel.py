import sys
import os
# Esto añade la carpeta raiz
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from Tools import Get_Resource_Path , Delete_Actual_Window , Save_New_Configurations_In_JSON_File , Read_Data_From_JSON , Check_Threads_Alive , Center_Window , Insert_Data_In_Log_File
from Calcs.Imports.Import_Data_From_Excel import Import_Excel_Using_Single_Range_Of_Cells
from Calcs.Imports.Import_Data_From_Excel import Import_Excel_Using_Multiple_Range_Of_Cells
from Window_Progress_Bar import W_Progress_Bar
from Exceptions.Exception_Warning import Raise_Warning

from tkinter import *
import os
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import pandas as pd # type: ignore
import threading
import openpyxl
from python_calamine import CalamineWorkbook

import time

def index_to_string(i):
    Letter = ''
    Temp = i
    while Temp >= 0:
        Letter = chr(Temp % 26 + 65) + Letter
        Temp = Temp // 26 - 1
    return Letter

class TreeviewFrame_Preview(ttk.Frame):
    def __init__(self , *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.hscrollbar = ttk.Scrollbar(self, orient=HORIZONTAL)
        self.vscrollbar = ttk.Scrollbar(self, orient=VERTICAL)
        self.treeview = ttk.Treeview(
            self,
            xscrollcommand=self.hscrollbar.set,
            yscrollcommand=self.vscrollbar.set
        )
        self.hscrollbar.config(command=self.treeview.xview)
        self.hscrollbar.pack(side=BOTTOM, fill=X)
        self.vscrollbar.config(command=self.treeview.yview)
        self.vscrollbar.pack(side=RIGHT, fill=Y)
        self.treeview.pack(fill="both" , expand=True)

        self.W_Import_Excel = None

        self.Total_Rows_In_Excel = None
        self.Total_Columns_In_Excel = None

        self.Excel_Dataframe = pd.DataFrame()
        self.First_One_Hundred_Data = None
        self.List_Number_Data_In_Row = None

        self.Error_In_Thread = False

    def Has_Rows(self):
        return len(self.treeview.get_children()) > 0
    
    def clear_table(self):
        if(self.Has_Rows()):
            for item in self.treeview.get_children():
                self.treeview.delete(item)
            self.treeview["columns"] = []

    def Display(self):
        self.place(x=40 , y=430)

    def Hidden(self):
        self.place_forget()

    def Fix_Invalid_Sheet_Number(self , Sheet_Number_Intvar):
        try:
            Sheet_Number_Intvar.set(Sheet_Number_Intvar.get() - 1)
            raise Raise_Warning(f"El numero de hoja {Sheet_Number_Intvar.get() + 1} no existe.")
        except Raise_Warning as e:
            self.Error_In_Thread = True
            messagebox.showwarning("Advertencia", f"{e}")
            Insert_Data_In_Log_File(e, "Advertencia", "Importacion de datos")

    def Function_Start_Thread(self , File_Path , Sheet_Number_Value , Sheet_Number_Intvar , Class_Progress_Bar , Function_Close_Thread = None):
        Thread_List = []

        Thread = threading.Thread(target= lambda: self.Get_Excel_File(File_Path , Sheet_Number_Value , Sheet_Number_Intvar))
        Thread_List.append(Thread)
        Thread.start()

        self.W_Import_Excel.after(500 , Check_Threads_Alive , Thread_List , self.W_Import_Excel , Class_Progress_Bar , Function_Close_Thread)

    def Function_Close_Thread(self):
        if(self.Error_In_Thread):
            self.Error_In_Thread = False
        else:
            self.Load_Excel_In_Preview()

    def Get_Excel_File(self , File_Path , Sheet_Number_Value , Sheet_Number_Intvar):
        try:
            if(File_Path):
                prev_load_excel = openpyxl.load_workbook(File_Path , read_only=True , data_only=True , keep_links=False)
                self.sheets = prev_load_excel.sheetnames

                if(Sheet_Number_Value > len(self.sheets)):
                    self.W_Import_Excel.after(550 , lambda: self.Fix_Invalid_Sheet_Number(Sheet_Number_Intvar))
                    return
                
                Idx_Sheet = Sheet_Number_Value - 1

                Sheet_Name = self.sheets[Idx_Sheet]
                One_Sheet = prev_load_excel[Sheet_Name]

                self.Total_Rows_In_Excel = One_Sheet.max_row
                self.Total_Columns_In_Excel = One_Sheet.max_column

                self.Get_Sheet_Data(File_Path , Idx_Sheet)

        except Raise_Warning as e:
            self.Error_In_Thread = True
            self.W_Import_Excel.after(0 , messagebox.showwarning("Advertencia" , f"{e}"))
            self.W_Import_Excel.after(30 , Insert_Data_In_Log_File(e , "Advertencia" , "Importacion de datos"))
            return
        except Exception as e:
            self.Error_In_Thread = True
            self.W_Import_Excel.after(0 , messagebox.showerror("Error" , f"{e}"))
            self.W_Import_Excel.after(30 , Insert_Data_In_Log_File(e , "Error" , "Importacion de datos"))
            return

    def Get_Sheet_Data(self , File_Path , Idx_Sheet):
        JSON_Settings_Data = Read_Data_From_JSON("import_excel_settings")

        Data_Excel = CalamineWorkbook.from_path(File_Path).get_sheet_by_index(Idx_Sheet).to_python(skip_empty_area=False)
        if(Data_Excel):
            self.Excel_Dataframe = pd.DataFrame(data=Data_Excel[1:] , columns=Data_Excel[0])
        else:
            self.Excel_Dataframe = pd.DataFrame(data=[""] , columns=["No hay datos"])

        self.First_One_Hundred_Data = self.Excel_Dataframe.head(JSON_Settings_Data["maximun_rows_to_display_in_preview"] - 1)
        self.List_Number_Data_In_Row = []
        Values_Recognized_as_Null = ["" , "NaN" , "None"]

        for col_pos in range(self.Excel_Dataframe.shape[1]):
            val = self.Excel_Dataframe.iloc[: , col_pos].to_list()
            counter_data_in_row = 0
            counter_void_places = 0
            for i , data in enumerate(val):
                if(not str(data).replace(" ","") in Values_Recognized_as_Null):
                    if(counter_void_places != 0):
                        counter_data_in_row += counter_void_places
                        counter_void_places = 0
                    
                    counter_data_in_row += 1
                else:
                    counter_void_places += 1
                
                if(counter_void_places > JSON_Settings_Data["void_tolerance"]):
                    break

            self.List_Number_Data_In_Row.append(counter_data_in_row + 1)
    
    def Load_Excel_In_Preview(self):
        self.clear_table()

        self.treeview["columns"] = []
        self.treeview["columns"] = ["N° fila/columna"] + [f"{i}" for i in range(len(self.First_One_Hundred_Data.columns))]
        Void_Space_In_Bottom_Preview = [""] + ["" for _ in range(len(self.First_One_Hundred_Data.columns))]

        self.treeview.heading("N° fila/columna", text="N° fila/columna")
        self.treeview.column("N° fila/columna" , anchor="center" , width=120 , stretch=False)
        List_With_All_Columns_Letters = []
        for i , col in enumerate(self.First_One_Hundred_Data.columns):
            Col_Letter = index_to_string(i)
            List_With_All_Columns_Letters.append(Col_Letter)

            self.treeview.heading(f"{i}" , text=Col_Letter)
            self.treeview.column(f"{i}" , anchor="center" , width=120 , stretch=False)

        val = tuple([1] + self.First_One_Hundred_Data.columns.tolist())
        self.treeview.insert("" , "end" , values=val)
        for (index, row) in self.First_One_Hundred_Data.iterrows():
            values = tuple([index + 2] + row.tolist())
            self.treeview.insert("", "end", values=values)

        self.treeview.insert("", "end", values=tuple(Void_Space_In_Bottom_Preview))
        
        Values_For_Bottom_Preview = [f"{col_letter}{row_count}" for col_letter , row_count in zip(List_With_All_Columns_Letters , self.List_Number_Data_In_Row)]
        self.treeview.insert("", "end", values=tuple(["Ultimo dato en:"] + Values_For_Bottom_Preview))

class Spinbox_With_Validation:
    def __init__(self , Root_Window , Max_Value , Min_Value , Increment_Value , Spinbox_Width , Value_Associed , **Place):
        self.Register_For_Spinbox = (Root_Window.register(self.Avoid_Unwanted_Values_In_Spinbox), '%P')
        self.Min_Value = Min_Value
        self.Max_Value = Max_Value
        
        self.Spinbox_In_App = Spinbox(Root_Window , textvariable=Value_Associed , from_=Min_Value , to=Max_Value , increment=Increment_Value , width=Spinbox_Width , font=("Courier New" , 13) , validate="all" , validatecommand=self.Register_For_Spinbox)
        self.Spinbox_In_App.place(x=Place["x"] , y=Place["y"])

    def Avoid_Unwanted_Values_In_Spinbox(self , Actual_Spinbox_Value):
        if(Actual_Spinbox_Value == ""):
            return True
        try:
            Number = int(Actual_Spinbox_Value)
            if("." in Actual_Spinbox_Value):
                return False
            return self.Min_Value <= Number <= self.Max_Value
        except ValueError:
            return False

    
def Select_File(W_Import_Excel , Path , Preview , Sheet_Number):
    Path_File = filedialog.askopenfilename(filetypes=[("Archivos Excel" , "*.xlsx")])
    if Path_File:
        if Path:
            Path.set("")
            Path.set(Path_File)
        else:
            Path.set(Path_File)

        Load_Excel_To_Preview(W_Import_Excel , Path , Sheet_Number , Preview)

def Load_Excel_To_Preview(W_Import_Excel , Path, Sheet_Number , Table_Preview_Data):
    if(Path):
        try:
            Class_Progress_Bar = W_Progress_Bar(W_Import_Excel)
            Class_Progress_Bar.Start_Progress_Bar("Cargando excel, esto podria tomar un momento...")

            if (not os.path.exists(Path.get())):
                Path.set("")
                raise Raise_Warning("El archivo Excel no existe en la ruta especificada.")

            if(isinstance(Sheet_Number.get() , float)):
                Sheet_Number.set(1)
                raise Raise_Warning("Numero de hoja no valido, solo valores enteros")
            Path_Value = Path.get()
            Sheet_Number_Value = Sheet_Number.get()

            Table_Preview_Data.Function_Start_Thread(Path_Value , Sheet_Number_Value , Sheet_Number , Class_Progress_Bar , Function_Close_Thread = Table_Preview_Data.Function_Close_Thread)
        except Raise_Warning as e:
            Class_Progress_Bar.Close_Progress_Bar()
            messagebox.showwarning("Advertencia" , f"{e}")
            Insert_Data_In_Log_File(e , "Advertencia" , "Importacion de datos")
        except Exception as e:
            Class_Progress_Bar.Close_Progress_Bar()
            messagebox.showerror("Error" , f"{e}")
            Insert_Data_In_Log_File(e , "Error" , "Importacion de datos")

def Call_Import_Classes(W_Import_Excel , Cell_Range , Table_Preview_Data , Source_Module_Name , Entry_Widget_For_W_Table_Frecuency , Value_For_Entry_Widget_W_Table_Frecuency , Imported_Data_From_Excel):
    try:
        """ 
            FALTA MODIFICAR ESTO PARA EVITAR RUNTIME ERROR CON THREADS
        """
        if(not Cell_Range.get()):
            raise Raise_Warning("No se ha ingresado un rango de celdas.")
        
        if(";" in Cell_Range.get()):
            Class_For_Import_Excel = Import_Excel_Using_Multiple_Range_Of_Cells(W_Import_Excel , Table_Preview_Data , Cell_Range.get() , Source_Module_Name , Entry_Widget_For_W_Table_Frecuency , Value_For_Entry_Widget_W_Table_Frecuency , Imported_Data_From_Excel)

            Class_For_Import_Excel.Process_Input_Data()
            Class_For_Import_Excel.Start_Thread_For_Import_Of_Data(Function_Close_Thread = Class_For_Import_Excel.Function_Close_Thread)

        elif(":" in Cell_Range.get()):
            Class_For_Import_Excel = Import_Excel_Using_Single_Range_Of_Cells(W_Import_Excel , Table_Preview_Data , Cell_Range.get() , Source_Module_Name , Entry_Widget_For_W_Table_Frecuency , Value_For_Entry_Widget_W_Table_Frecuency , Imported_Data_From_Excel)

            Class_For_Import_Excel.Process_Input_Data()
            Class_For_Import_Excel.Start_Thread_For_Import_Of_Data(Function_Close_Thread = Class_For_Import_Excel.Function_Close_Thread)
        else:
            raise Raise_Warning("El rango de celdas ingresado es invalido.")

    except (FileNotFoundError , Raise_Warning) as e:
        messagebox.showwarning("Advertencia" , f"{e}")
        Insert_Data_In_Log_File(e , "Advertencia" , "Importacion de datos")
    except Exception as e:
        messagebox.showerror("Error" , f"{e}")
        Insert_Data_In_Log_File(e , "Error" , "Importacion de datos")

def Create_Window_Import_Configuration(W_Import_Excel=None):
    def Close_Settings_Window():
        try:
            Save_New_Configurations_In_JSON_File("import_excel_settings" , void_tolerance=Void_Tolerance_Number.get() , maximun_rows_to_display_in_preview=Number_Rows_To_Display.get() , import_matrix_data=Import_Data_Matrix.get())
        except Exception as e:
            messagebox.showerror("Error" , "Algo salio mal al guardar la configuracion.\nAsegurese que todos los datos esten bien escritos.")
            Insert_Data_In_Log_File("Algo salio mal al guardar la configuracion. Asegurese que todos los datos esten bien escritos." , "Error" , "Importacion de datos" , e)
        else:
            Delete_Actual_Window(W_Import_Excel , W_Import_Configuration)
            Insert_Data_In_Log_File("Datos guardados correctamente en el archivo de configuracion" , "Operacion exitosa" , "Importacion de datos")
    
    JSON_Settings_Data = Read_Data_From_JSON("import_excel_settings")
    
    W_Import_Configuration = Toplevel(W_Import_Excel)

    W_Import_Configuration.grab_set()
    # W_Import_Configuration.geometry("500x200+500+350")
    Center_Window(W_Import_Configuration , 500 , 200)
    W_Import_Configuration.title("Configurar Importacion")
    W_Import_Configuration.config(bg="#d1e7d2")
    Icon = PhotoImage(file=Get_Resource_Path("Images/icon.png"))
    W_Import_Configuration.iconphoto(False , Icon)
    W_Import_Configuration.protocol("WM_DELETE_WINDOW" , Close_Settings_Window)
    W_Import_Configuration.lift()

    Void_Tolerance_Number = IntVar(W_Import_Configuration)
    Number_Rows_To_Display = IntVar(W_Import_Configuration)
    Import_Data_Matrix = BooleanVar(W_Import_Configuration)

    Label_Input_Void_Tolerance = Label(W_Import_Configuration , text="Tolerancia de celdas vacias (0 - 25):" , font=("Times New Roman" , 12) , bg="#d1e7d2")
    Label_Input_Void_Tolerance.place(x=20 , y=20)
    Input_Void_Tolerance = Spinbox_With_Validation(W_Import_Configuration , 25 , 0 , 1 , 3 , Void_Tolerance_Number , x=400 , y=20)
    Void_Tolerance_Number.set(JSON_Settings_Data["void_tolerance"])

    Label_Input_Number_Rows_To_Display = Label(W_Import_Configuration , text="Filas en la previsualizacion (0 - 2000)" , font=("Times New Roman" , 13) , bg="#d1e7d2")
    Label_Input_Number_Rows_To_Display.place(x=20 , y=60)
    Input_Number_Rows_To_Display = Spinbox_With_Validation(W_Import_Configuration , 1000 , 10 , 10 , 5 , Number_Rows_To_Display , x=400 , y=60)
    Number_Rows_To_Display.set(JSON_Settings_Data["maximun_rows_to_display_in_preview"])

    #Checkbox_Import_Data_Matrix = Checkbutton(W_Import_Configuration , text="Importar matriz de datos" , font=("Times New Roman" , 13) , bg="#d1e7d2" , variable=Import_Data_Matrix)
    #Checkbox_Import_Data_Matrix.place(x=20 , y=100)
    #Import_Data_Matrix.set(JSON_Settings_Data["import_matrix_data"])

    W_Import_Configuration.resizable(False , False)
    W_Import_Excel.wait_window(W_Import_Configuration)

    W_Import_Configuration.mainloop()

def Create_Window_Import_Excel(Father_Window , Value_For_Entry_Widget_W_Table_Frecuency , Entry_Widget_For_W_Table_Frecuency , Imported_Data_From_Excel , Source_Module_Name):
    W_Import_Excel = Toplevel(Father_Window)

    Icon = PhotoImage(file=Get_Resource_Path("Images/icon.png"))

    W_Import_Excel.grab_set()
    Center_Window(W_Import_Excel , 800 , 550)

    W_Import_Excel.title("Seleccionar Archivo")
    W_Import_Excel.config(bg="#d1e7d2")
    W_Import_Excel.iconphoto(False , Icon)
    W_Import_Excel.protocol("WM_DELETE_WINDOW" , lambda: Delete_Actual_Window(Father_Window , W_Import_Excel))
    
    Path = StringVar(W_Import_Excel)
    Cell_Range = StringVar(W_Import_Excel)
    Sheet_Number = IntVar(W_Import_Excel)

    Btn_Configuracion = Button(W_Import_Excel , text="\u2699" , font=("Segoe UI Emoji", 9) , bg="#d1e7d2" , command= lambda: Create_Window_Import_Configuration(W_Import_Excel))
    Btn_Configuracion.place(x=20 , y=312)

    Text_Input_Path_File = Label(W_Import_Excel , text="Ingrese la ruta del archivo: " , bg="#d1e7d2" , font=("Times New Roman" , 13))
    Text_Input_Path_File.place(x=20 , y=340)
    Path_File = Entry(W_Import_Excel , font=("Courier New" , 13) , textvariable=Path , width=55 , state="readonly")
    Path_File.place(x=210 , y=340)
    Btn_Select_File = Button(W_Import_Excel , text="Examinar" , font=("Times New Roman" , 13) , command= lambda: Select_File(W_Import_Excel , Path , Table_Preview_Data , Sheet_Number) , width=10 , bg="#ffe3d4")
    Btn_Select_File.place(x=50 , y=370)

    Text_Input_Sheet_Number = Label(W_Import_Excel , text="Numero de Hoja: " , bg="#d1e7d2" , font=("Times New Roman" , 13))
    Text_Input_Sheet_Number.place(x=20 , y=410)
    Input_Sheet_Number = Spinbox(W_Import_Excel , font=("Courier New" , 13) , textvariable=Sheet_Number , from_=1 , to=100 , width=4 , state="readonly" , command= lambda: Load_Excel_To_Preview(W_Import_Excel , Path , Sheet_Number , Table_Preview_Data))
    Input_Sheet_Number.place(x=210 , y=410)

    Text_Input_Cells_Range = Label(W_Import_Excel , text="Ingrese el rango de celdas:" , bg="#d1e7d2" , font=("Times New Roman" , 13))
    Text_Input_Cells_Range.place(x=20 , y=450)
    Cells_Range = Entry(W_Import_Excel , font=("Courier New" , 13) , textvariable=Cell_Range , width=55)
    Cells_Range.place(x=210 , y=460)
    Cells_Range.focus()

    Table_Preview_Data = TreeviewFrame_Preview(W_Import_Excel)
    Table_Preview_Data.W_Import_Excel = W_Import_Excel
    Table_Preview_Data.pack(fill=BOTH)
    Table_Preview_Data.treeview.config(height=9)
    Table_Preview_Data.treeview.config(columns=("1", "2" ,"3", "4", "5" , "6") , show="headings")
    Table_Preview_Data.treeview.heading("1" , text="fila/columna")
    Table_Preview_Data.treeview.heading("2" , text="A")
    Table_Preview_Data.treeview.heading("3" , text="B")
    Table_Preview_Data.treeview.heading("4" , text="C")
    Table_Preview_Data.treeview.heading("5" , text="D")
    Table_Preview_Data.treeview.heading("6" , text="E")
    for a in range(1 , 7):
        Table_Preview_Data.treeview.column(f"{a}" , anchor="center" , width=106 , stretch=True)

    Btn_Process_Data = Button(W_Import_Excel , text="Importar Datos" , font=("Times New Roman" , 13) , width=25 , bg="#ffe3d4" , command=lambda: Call_Import_Classes(W_Import_Excel , Cell_Range , Table_Preview_Data , Source_Module_Name , Entry_Widget_For_W_Table_Frecuency , Value_For_Entry_Widget_W_Table_Frecuency , Imported_Data_From_Excel))
    Btn_Process_Data.pack(side=BOTTOM)

    W_Import_Excel.resizable(False,False)
    Father_Window.wait_window(W_Import_Excel)

    W_Import_Excel.mainloop()

if __name__ == "__main__":
    Create_Window_Import_Configuration()