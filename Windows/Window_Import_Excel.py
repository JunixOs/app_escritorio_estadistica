from Window_Progress_Bar import W_Progress_Bar

from tkinter import *
import os
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import pandas as pd # type: ignore
import openpyxl
from openpyxl.utils import column_index_from_string
import threading
import re

class TreeviewFrame(ttk.Frame):
    def __init__(self , *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.hscrollbar = ttk.Scrollbar(self, orient=HORIZONTAL)
        self.vscrollbar = ttk.Scrollbar(self, orient=VERTICAL)
        self.treeview = ttk.Treeview(
            self,
            xscrollcommand=self.hscrollbar.set,
            yscrollcommand=self.vscrollbar.set
        )
        self.hscrollbar.config(command=self.treeview.xview)
        self.hscrollbar.pack(side=BOTTOM, fill=X)
        self.vscrollbar.config(command=self.treeview.yview)
        self.vscrollbar.pack(side=RIGHT, fill=Y)
        self.treeview.pack(fill="both" , expand=True)

        self.Progress_Bar = None
        self.Root_Window = None

    def Has_Rows(self):
        return len(self.treeview.get_children()) > 0
    
    def clear_table(self):
        if(self.Has_Rows()):
            for item in self.treeview.get_children():
                self.treeview.delete(item)
            self.treeview["columns"] = []

    def Display(self):
        self.place(x=40 , y=430)

    def Hidden(self):
        self.place_forget()

    def Insert_Imported_Data_To_Preview(self, data , start_row , end_row):
        # encabezados
        self.treeview.delete(*self.treeview.get_children())

        data = data.dropna(axis=1, how='all')
        self.treeview["columns"] = []
        self.treeview["columns"] = ["fila"] + data.columns.tolist()

        self.treeview.heading("fila" , text="N° fila")
        self.treeview.column("fila" , anchor="center" , width=120 , stretch=False)
        for col in data.columns:
            self.treeview.heading(col , text=col)
            self.treeview.column(col , anchor="center" , width=120 , stretch=False)

        Dot_Text = tuple(["......."] for _ in range(0 , len(self.treeview["columns"])))

        # Insertar los datos fila por fila
        if(end_row - start_row + 1 >= 100):
            for (index, row) in (data.head().iterrows()):
                    values = tuple([index + 2] + row.tolist())
                    self.treeview.insert("" , "end" , values=values)
            for i in range(0 , 3):
                self.treeview.insert("" , "end" , values=Dot_Text)
            for (index, row) in (data.tail().iterrows()):
                    values = tuple([index + 2] + row.tolist())
                    self.treeview.insert("" , "end" , values=values)
        else:
            for (index, row) in (data.iterrows()):
                    values = tuple([index + 1] + row.tolist())
                    self.treeview.insert("" , "end" , values=values)

        self.Progress_Bar.Close_Progress_Bar()

        messagebox.showinfo("Success" , "Datos procesados con exito.\nYa puede salir de la ventana de importacion.")

    def Load_Excel_File(self , File_Path , Sheet_Number):
        try:
            if(File_Path):
                self.data = pd.ExcelFile(File_Path)
                self.sheets = self.data.sheet_names

                if(Sheet_Number.get() > len(self.sheets)):
                    Sheet_Number.set(Sheet_Number.get() - len(self.sheets))
                    raise Exception(f"El numero de hoja {Sheet_Number.get() + len(self.sheets)} no existe.")
                
                Sheet_N = Sheet_Number.get() - 1

                self.Load_Sheet_Data(Sheet_N)
        except Exception as e:
            self.Progress_Bar.Close_Progress_Bar()
            messagebox.showerror("Error" , f"{e}")

    def Load_Sheet_Data(self , Sheet_Number):
        self.treeview.delete(*self.treeview.get_children())

        data = self.data.parse(sheet_name=Sheet_Number)

        data = data.head(50)

        self.treeview["columns"] = []
        self.treeview["columns"] = ["N° fila/columna"] + [f"{i}" for i in range(len(data.columns))]

        self.treeview.heading("N° fila/columna", text="N° fila/columna")
        self.treeview.column("N° fila/columna" , anchor="center" , width=120 , stretch=False)
        for i , col in enumerate(data.columns):
            Col_Letter = ''
            Temp = i
            while Temp >= 0:
                Col_Letter = chr(Temp % 26 + 65) + Col_Letter
                Temp = Temp // 26 - 1
            self.treeview.heading(f"{i}" , text=Col_Letter)
            self.treeview.column(f"{i}" , anchor="center" , width=120 , stretch=False)

        # Insertar los datos en el Treeview
        for (index, row) in data.iterrows():
            values = tuple([index + 2] + row.tolist())
            self.treeview.insert("", "end", values=values)

        self.Progress_Bar.Close_Progress_Bar()
    
def Select_File(Path , Preview , Sheet_Number):
    Path_File = filedialog.askopenfilename(filetypes=[("Archivos Excel" , "*.xlsx")])
    if Path_File:
        if Path:
            Path.set("")
            Path.set(Path_File)
        else:
            Path.set(Path_File)

        Load_Excel_To_Preview(Path , Sheet_Number , Preview)

def Load_Excel_To_Preview(Path, Sheet_Number , Preview):
    if(Path):
        try:
            Preview.Progress_Bar.Start_Progress_Bar("Cargando excel, esto podria tomar un momento...")

            if (not os.path.exists(Path.get())):
                Path.set("")
                raise Exception("El archivo Excel no existe en la ruta especificada.")

            if(isinstance(Sheet_Number.get() , float)):
                Sheet_Number.set(1)
                raise Exception("Numero de hoja no valido, solo valores enteros")

            threading.Thread(target= lambda: Preview.Load_Excel_File(Path.get() , Sheet_Number)).start()
        except Exception as e:
            Preview.Progress_Bar.Close_Progress_Bar()
            messagebox.showerror("Error" , f"{e}")

def Import_Data_From_Single_Column(File_Path , Widget_Sheet_Number , column , start_row , end_row , Preview , Data_From_Widget_Entry , Data_From_Single_Column , Input_Data):
    try:

        Load_Excel = openpyxl.load_workbook(File_Path.get() , read_only=True)
        Sheet_Number = Widget_Sheet_Number.get()
        Sheet_Number -= 1

        Sheet_Name = Load_Excel.sheetnames[Sheet_Number]
        Sheet = Load_Excel[Sheet_Name]

        total_rows = Sheet.max_row
        total_columns = Sheet.max_column
        column_index = column_index_from_string(column)

        if(end_row > total_rows):
            raise Exception("Se intento acceder a una fila no valida, intente nuevamente.")
        elif(column_index > total_columns):
            raise Exception("Se intento acceder a una columna no valida, intente nuevamente.")


        Excel = pd.read_excel(File_Path.get() , sheet_name=Sheet_Number , engine="openpyxl" , usecols=f"{column}:{column}" , nrows=end_row + 10)
        if("Unnamed" in Excel.columns):
            raise Exception("Se intento importar datos sin un encabezado adecuado. Por favor, coloque un nombre adecuado a los datos y coloquelos en la primera fila.")
        
        if(start_row == 1):
            data = Excel.iloc[start_row-1:end_row-1]
        else:
            data = Excel.iloc[start_row-2:end_row-1]

        data.dropna()

        if data.isnull().all().all():
            raise Exception("Los datos seleccionados están vacíos o contienen solo valores nulos. Por favor, intente con otra columna")
        
        if data.isnull().any().any():
            raise Exception("Los datos seleccionados contienen algun valor nulo. Por favor, revise si los datos tienen un formato adecuado o si el rango de celdas que ingreso cubre solamente los datos a importar y ninguna celda mas.")
        if("Unnamed" in Excel.columns[0]):
            raise Exception("Se intento importar datos sin un encabezado adecuado. Por favor, coloque un nombre adecuado a los datos y coloquelos en la primera fila.")
        
        Preview.clear_table()
        Preview.Insert_Imported_Data_To_Preview(data , start_row , end_row)

        if(Data_From_Widget_Entry):
            Data_From_Widget_Entry.set("")
        if(Data_From_Single_Column):
            Data_From_Single_Column.clear()
        Data_From_Single_Column[f"{Excel.columns[0]}"] = [value[0] for value in data.values]
        
        Load_Excel.close()

        Data_From_Widget_Entry.set(f"Columna Importada: {Excel.columns[0]}")

    except (FileNotFoundError , Exception) as e:
        Preview.Progress_Bar.Close_Progress_Bar()
        messagebox.showerror("Error" , f"{e}")
    else:
        Input_Data.config(state="disabled")

def Import_Data_From_Multiple_Columns(File_Path , Widget_Sheet_Number , start_column , end_column , start_row , end_row , Ranges , Preview , Data_From_Widget_Entry , Data_From_Multiple_Columns , Input_Data):
    try:
        Load_Excel = openpyxl.load_workbook(File_Path.get() , read_only=True)
        Sheet_Number = Widget_Sheet_Number.get()
        Sheet_Number -= 1

        Sheet_Name = Load_Excel.sheetnames[Sheet_Number]
        Sheet = Load_Excel[Sheet_Name]

        total_rows = Sheet.max_row
        total_columns = Sheet.max_column
        start_column_index = column_index_from_string(start_column)
        end_column_index = column_index_from_string(end_column)

        if(Ranges):
            Columns = []
            Rows = []
            for r in (Ranges["Rows"]):
                if(r[0] > total_rows or r[1] > total_rows):
                    raise Exception("Se intento acceder a una fila no valida, intente nuevamente.")
                Rows.append(r[0])
                Rows.append(r[1])

            for c in (Ranges["Columns"]):
                col_start = column_index_from_string(c[0])
                col_end = column_index_from_string(c[1])
                if(col_start > total_columns or col_end > total_columns):
                    raise Exception("Se intento acceder a una columna no valida, intente nuevamente.")
                Columns.append(c[0])
                Columns.append(c[1])
            Unique_Columns = list(set(Columns))
            String_Columns = ",".join(Unique_Columns)
    
            Excel = pd.read_excel(File_Path.get() , sheet_name=Sheet_Number , engine="openpyxl" , usecols=String_Columns , nrows=max(Rows) + 10)
            if("Unnamed" in Excel.columns):
                raise Exception("Se intento importar datos sin un encabezado adecuado. Por favor, coloque un nombre adecuado a los datos y coloquelos en la primera fila.")
            
            Columns_Name = Excel.columns
            Concat_Columns = []

            for i in range(0 , len(Ranges["Rows"])):
                col = list(set(Ranges["Columns"][i]))
                if(len(col) == 1):
                    col = column_index_from_string(col[0]) - 1
                else:
                    col = [column_index_from_string(c) - 1 for c in col]
                
                if(Ranges["Rows"][i][0] == 1):
                    column_i = Excel.iloc[Ranges["Rows"][i][0]-1:Ranges["Rows"][i][1]-1 , i]
                else:
                    column_i = Excel.iloc[Ranges["Rows"][i][0]-2:Ranges["Rows"][i][1]-1 , i]
                Concat_Columns.append(column_i)

            data = pd.concat(Concat_Columns , axis=1 , ignore_index=True)
            """ PERMITIR LA IMPORTACION DE CELDAS COMO C1:C1001;E1:F1001 """
            data.columns = Columns_Name

        else:
            if(start_row > total_rows or end_row > total_rows):
                raise Exception("Se intento acceder a una fila no valida, intente nuevamente.")
            elif(start_column_index > total_columns or end_column_index > total_columns):
                raise Exception("Se intento acceder a una columna no valida, intente nuevamente.")

            Excel = pd.read_excel(File_Path.get() , sheet_name=Sheet_Number , engine="openpyxl" , usecols=f"{start_column}:{end_column}" , nrows=end_row + 10)
            if("Unnamed" in Excel.columns):
                raise Exception("Se intento importar datos sin un encabezado adecuado. Por favor, coloque un nombre adecuado a los datos y coloquelos en la primera fila.")
            if(start_row == 1):
                data = Excel.iloc[start_row-1:end_row-1]
            else:
                data = Excel.iloc[start_row-2:end_row-1]

        if data.isnull().all().all():
            raise Exception("Los datos seleccionados están vacíos o contienen solo valores nulos.")
        
        if data.isnull().any().any():
            raise Exception("Los datos seleccionados contienen algun valor nulo. Por favor, revise si los datos tienen un formato adecuado.")

        Preview.clear_table()
        Preview.Insert_Imported_Data_To_Preview(data , start_row , end_row)
        
        if(Data_From_Widget_Entry):
            Data_From_Widget_Entry.set("")
        if(Data_From_Multiple_Columns):
            Data_From_Multiple_Columns.clear()

        text = "columnas importadas: "
        for Column in Excel.columns:
            Data_From_Multiple_Columns[Column] = [value for value in data[Column].dropna()]
            text = text + Column + "  "

        Load_Excel.close()

        Data_From_Widget_Entry.set(text)

    except (FileNotFoundError , Exception) as e:
        Preview.Progress_Bar.Close_Progress_Bar()
        messagebox.showerror("Error" , f"{e}")
    else:
        Input_Data.config(state="disabled")

def Validate_Range_Of_Cells(Range):
    Range = Range.upper()
    return re.match(r"([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)" , Range.strip())

def Process_File_Data(File_Path , Widget_Sheet_Number , Cell_Range , Preview , Data_From_Widget_Entry , Input_Data , Data_From_Single_Column , Data_From_Multiple_Columns):
    """ Separar en diferentes ventanas, uno para importar de un .xlsx y otro para importar de un .txt """
    try:
        Preview.Progress_Bar.Start_Progress_Bar()

        Dict_Cell_Ranges = {
            "Columns": [],
            "Rows": [],
        }
        Non_Consecutive_Columns = False

        Sheet_Number = Widget_Sheet_Number.get()
        if(not File_Path.get()):
            raise Exception("No se ha ingresado la ruta del archivo.")
        
        if (not os.path.exists(File_Path.get())):
            raise Exception("El archivo Excel no existe en la ruta especificada.")

        if(isinstance(Sheet_Number , float)):
            raise Exception("Numero de hoja no valido, solo valores enteros")
        
        Unload_Excel = pd.ExcelFile(f"{File_Path.get()}")
        Sheets = Unload_Excel.sheet_names
        if(Sheet_Number > len(Sheets)):
            raise Exception("No existe el numero de hoja especificado")
        if(not Cell_Range.get()):
            raise Exception("No se ha ingresado un rango de celdas.")
        
        if(";" in Cell_Range.get()):
            Ranges = Cell_Range.get().split(";")
            for ran in Ranges:
                ran = Validate_Range_Of_Cells(ran)
                if(not ran):
                    raise Exception("El rango de celdas ingresado es invalido.")

                column_start, start_row, column_end, end_row = ran.groups()

                start_row = int(start_row)
                end_row = int(end_row)

                if column_start == column_end and start_row == end_row:
                    raise Exception("Una de las selecciones corresponde una celda individual, este no es un rango válido.")
                elif(start_row > end_row):
                    start_row , end_row = end_row , start_row
                elif(column_start > column_end):
                    column_start , column_end = column_end , column_start

                Dict_Cell_Ranges["Columns"].append([column_start , column_end])

                Dict_Cell_Ranges["Rows"].append([start_row , end_row])

            Non_Consecutive_Columns = True

        elif(":" in Cell_Range.get()):
            Range = Validate_Range_Of_Cells(Cell_Range.get())
            if(not Range):
                raise Exception("El rango de celdas ingresado es invalido.")
            column_start, start_row, column_end, end_row = Range.groups()

            start_row = int(start_row)
            end_row = int(end_row)

            if (column_start == column_end and start_row == end_row):
                raise Exception("Seleccionaste una celda individual, no es un rango válido.")
            elif(start_row > end_row):
                start_row , end_row = end_row , start_row
            elif(column_start > column_end):
                column_start , column_end = column_end , column_start
            Dict_Cell_Ranges = {}
        else:
            raise Exception("El rango de celdas ingresado es invalido.")
        
        match(column_start != column_end or Non_Consecutive_Columns):
            case True:
                threading.Thread(target= lambda: Import_Data_From_Multiple_Columns(File_Path , Widget_Sheet_Number , column_start , column_end , start_row , end_row , Dict_Cell_Ranges , Preview , Data_From_Widget_Entry , Data_From_Multiple_Columns , Input_Data)).start()
            case False:
                threading.Thread(target= lambda: Import_Data_From_Single_Column(File_Path , Widget_Sheet_Number , column_start , start_row , end_row , Preview , Data_From_Widget_Entry , Data_From_Single_Column , Input_Data)).start()
            case _:
                raise Exception("Hubo un error al realizar la importacion.")

    except (FileNotFoundError , Exception) as e:
        Preview.Progress_Bar.Close_Progress_Bar()
        messagebox.showerror("Error" , f"{e}")

def Create_Window_Import_Excel(Father_Window , Data_From_Widget_Entry , Input_Data , Data_From_Single_Column , Data_From_Multiple_Columns):
    def Back():
        for widget in W_Import_Excel.winfo_children():
            widget.destroy()
        W_Import_Excel.grab_release()
        W_Import_Excel.quit()
        W_Import_Excel.destroy()

        Father_Window.lift()
    if __name__ == "__main__":
        W_Import_Excel = Tk()
    else:
        W_Import_Excel = Toplevel(Father_Window)

    Icon = PhotoImage(file="Images/icon.png")

    W_Import_Excel.grab_set()
    W_Import_Excel.geometry("800x550+350+170")
    W_Import_Excel.title("Seleccionar Archivo")
    W_Import_Excel.config(bg="#d1e7d2")
    W_Import_Excel.iconphoto(False , Icon)
    W_Import_Excel.protocol("WM_DELETE_WINDOW" , Back)
    
    Path = StringVar(W_Import_Excel)
    Cell_Range = StringVar(W_Import_Excel)
    Sheet_Number = IntVar(W_Import_Excel)

    Progress_Bar = W_Progress_Bar(W_Import_Excel)

    Text_Input_Path_File = Label(W_Import_Excel , text="Ingrese la ruta del archivo: " , bg="#d1e7d2" , font=("Times New Roman" , 12))
    Text_Input_Path_File.place(x=20 , y=340)
    Path_File = Entry(W_Import_Excel , font=("Courier New" , 11) , textvariable=Path , width=55 , state="readonly")
    Path_File.place(x=210 , y=340)
    Btn_Select_File = Button(W_Import_Excel , text="Examinar" , font=("Times New Roman" , 13) , command= lambda: Select_File(Path , Table_Preview_Data , Sheet_Number) , width=10 , bg="#ffe3d4")
    Btn_Select_File.place(x=50 , y=370)

    Text_Input_Sheet_Number = Label(W_Import_Excel , text="Numero de Hoja: " , bg="#d1e7d2" , font=("Times New Roman" , 13))
    Text_Input_Sheet_Number.place(x=20 , y=410)
    Input_Sheet_Number = Spinbox(W_Import_Excel , font=("Courier New" , 13) , textvariable=Sheet_Number , from_=1 , to=100 , width=4 , state="readonly" , command= lambda: Load_Excel_To_Preview(Path , Sheet_Number , Table_Preview_Data))
    Input_Sheet_Number.place(x=210 , y=410)

    Text_Input_Cells_Range = Label(W_Import_Excel , text="Ingrese el rango de celdas:\nSolo los datos" , bg="#d1e7d2" , font=("Times New Roman" , 13))
    Text_Input_Cells_Range.place(x=20 , y=440)
    Cells_Range = Entry(W_Import_Excel , font=("Courier New" , 13) , textvariable=Cell_Range , width=55)
    Cells_Range.place(x=210 , y=440)
    Cells_Range.focus()

    Table_Preview_Data = TreeviewFrame(W_Import_Excel)
    Table_Preview_Data.Progress_Bar = Progress_Bar
    Table_Preview_Data.Root_Window = W_Import_Excel
    Table_Preview_Data.pack(fill=BOTH)
    Table_Preview_Data.treeview.config(height=13)
    Table_Preview_Data.treeview.config(columns=("1", "2" ,"3", "4", "5" , "6") , show="headings")
    Table_Preview_Data.treeview.heading("1" , text="fila/columna")
    Table_Preview_Data.treeview.heading("2" , text="A")
    Table_Preview_Data.treeview.heading("3" , text="B")
    Table_Preview_Data.treeview.heading("4" , text="C")
    Table_Preview_Data.treeview.heading("5" , text="D")
    Table_Preview_Data.treeview.heading("6" , text="E")
    for a in range(1 , 7):
        Table_Preview_Data.treeview.column(f"{a}" , anchor="center" , width=106 , stretch=True)

    Btn_Process_Data = Button(W_Import_Excel , text="Procesar Archivo" , font=("Times New Roman" , 13) , width=25 , bg="#ffe3d4" , command=lambda: Process_File_Data(Path , Sheet_Number , Cell_Range , Table_Preview_Data , Data_From_Widget_Entry , Input_Data , Data_From_Single_Column , Data_From_Multiple_Columns))
    Btn_Process_Data.pack(side=BOTTOM)

    W_Import_Excel.resizable(False,False)
    W_Import_Excel.mainloop()

if __name__ == "__main__":
    Create_Window_Import_Excel(None , "" , "" , "" , {})