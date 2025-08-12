import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from Tools import Insert_Data_In_Log_File , Get_Detailed_Info_About_Error , Read_Data_From_JSON
from Exceptions.Exception_Warning import Raise_Warning

from tkinter import *
from tkinter import messagebox
import pandas as pd # type: ignore
import openpyxl
from python_calamine import CalamineWorkbook
import re


def index_to_string(i):
    Letter = ''
    Temp = i
    while Temp >= 0:
        Letter = chr(Temp % 26 + 65) + Letter
        Temp = Temp // 26 - 1
    return Letter

def string_to_index(s):
    result = 0
    for char in s:
        result = result * 26 + (ord(char.upper()) - 65 + 1)
    return result - 1


class Validator:
    def Validate_Format_For_Each_Range_Cells(self , One_Range_Cells):
        One_Range_Cells = One_Range_Cells.upper()
        One_Range_Cells = re.match(r"([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)" , One_Range_Cells.strip())
        if(not One_Range_Cells):
            raise Raise_Warning("El rango de celdas ingresado es incorrecto.")
        return One_Range_Cells.groups()

    def Validate_Order_For_Each_Range_Cells(self , Start_Row , End_Row , Start_Column , End_Column):
        if(Start_Column == End_Column and Start_Row == End_Row):
            raise Raise_Warning("No se puede seleccionar una sola celda.")
        elif(Start_Row > End_Row):
            Start_Row , End_Row = End_Row , Start_Row
        elif(Start_Column > End_Column):
            Start_Column , End_Column = End_Column , Start_Column
        return Start_Row , End_Row , Start_Column , End_Column
    
    def Validate_Order_For_All_Range_Cells(self , Ranges_Of_Cells):
        Dict_Range_Cells = {}
        Order_For_Range_Cells = []
        for range_cells in Ranges_Of_Cells:
            if(range_cells[0].isdigit()):
                raise Raise_Warning("El rango de celdas ingresado es incorrecto.")
            
            Dict_Range_Cells[f"{string_to_index(range_cells[0])}"] = range_cells
            Order_For_Range_Cells.append(string_to_index(range_cells[0]))
        
        Order_For_Range_Cells.sort()
        return [Dict_Range_Cells[f"{idx}"] for idx in Order_For_Range_Cells]

    def Validate_For_Avoid_Repeated_Range_Cells(self , Ranges_Of_Cells):
        for range_cell in Ranges_Of_Cells:
            if(Ranges_Of_Cells.count(range_cell) > 1):
                raise Raise_Warning("No se puede importar el mismo rango de celdas dos veces.")

    def Validate_Row_Limit_Excel(self , Input_Start_Row , Input_End_Row , Limit_Excel_End_Rows):
        if(Input_Start_Row > Limit_Excel_End_Rows or Input_End_Row > Limit_Excel_End_Rows):
            raise Raise_Warning("Se intento acceder a una fila no valida, fuera de alcance o sin nungun dato. Intente nuevamente.")
        
    def Validate_Column_Limit_Excel(self , Input_Start_Column , Input_End_Column , Limit_Excel_End_Columns):
        Input_Start_Column = string_to_index(Input_Start_Column)
        Input_End_Column = string_to_index(Input_End_Column)
        
        if(Input_Start_Column > Limit_Excel_End_Columns or Input_End_Column > Limit_Excel_End_Columns):
            raise Raise_Warning("Se intento acceder a una columna no valida, fuera de alcance o sin nungun dato. Intente nuevamente.")
        
    def Validate_Data_Imported_Is_Null(self , Imported_Data , Validate_All_Imported_Data = True):
        if(Validate_All_Imported_Data):
            if(Imported_Data.isnull().all().all()):
                raise Raise_Warning("Los datos seleccionados están vacíos o contienen solo valores nulos.")
            if(Imported_Data.isnull().any().any()):
                raise Raise_Warning("Los datos seleccionados contienen algun valor nulo. Por favor, revise si los datos tienen un formato adecuado.")
        else:
            if(Imported_Data.isnull().all()):
                raise Raise_Warning("Los datos seleccionados están vacíos o contienen solo valores nulos.")
            if(Imported_Data.isnull().any()):
                raise Raise_Warning("Los datos seleccionados contienen algun valor nulo. Por favor, revise si los datos tienen un formato adecuado.")


class Loader_Of_Selected_Data_In_Table:
    def __init__(self):
        self.Table_For_Show_Selected_Data = None

        self.Entry_Widget_For_W_Table_Frecuency = None
        self.Value_For_Entry_Widget_W_Table_Frecuency = None

        self.Entry_Widgets_For_Venn_Diagram = None
        self.Values_For_Entry_Widgets_Venn_Diagram = None

        self.Imported_Data_From_Excel_For_Calcs = {}
        self.Import_Multiple_Columns = False
        self.Imported_Data = None
        self.Imported_Column_Names = None
        self.Start_Row = None
        self.End_Row = None

    def Load_Selected_Data_In_Table(self):
        self.Imported_Data = self.Imported_Data.dropna(axis=1, how='all')
        Titles_For_Columns = ["N° fila"] + self.Imported_Data.columns.tolist()

        self.Table_For_Show_Selected_Data.Modify_Number_Of_Columns(len(Titles_For_Columns) , Titles_For_Columns , 100)

        Data_To_Display = []

        Dot_Text = tuple(["......."] for _ in range(0 , len(Titles_For_Columns)))
        N_Imported_Rows_For_Column = self.Imported_Data.count().tolist()
        Total_Row_Text = tuple(["Datos Importados:"] + N_Imported_Rows_For_Column)

        if(self.End_Row - self.Start_Row + 1 >= 100):
            for (index, row) in (self.Imported_Data.head(20).iterrows()):
                Data_To_Display.append(tuple([index + 2] + row.tolist()))
            for _ in range(0 , 3):
                Data_To_Display.append(Dot_Text)
            for (index, row) in (self.Imported_Data.tail(10).iterrows()):
                Data_To_Display.append(tuple([index + 2] + row.tolist()))
        else:
            for (index, row) in (self.Imported_Data.iterrows()):
                Data_To_Display.append(tuple([index + 1] + row.tolist()))
        
        self.Table_For_Show_Selected_Data.Insert_Data(Data_To_Display , Total_Row_Text)

    def Manage_Load_For_Module_Table_Of_Frecuency(self):
        if(self.Value_For_Entry_Widget_W_Table_Frecuency.get()):
            self.Value_For_Entry_Widget_W_Table_Frecuency.set("")
        if(self.Imported_Data_From_Excel_For_Calcs):
            self.Imported_Data_From_Excel_For_Calcs.clear()

        match(self.Import_Multiple_Columns):
            case True:
                text = "columnas importadas: "

                for column in self.Imported_Column_Names:
                    self.Imported_Data_From_Excel_For_Calcs[column] = [value for value in self.Imported_Data[column].dropna()]
                    text = text + column + "  "

                self.Entry_Widget_For_W_Table_Frecuency.config(state="disabled")
                self.Value_For_Entry_Widget_W_Table_Frecuency.set(text)

                self.Load_Selected_Data_In_Table()

            case False:
                self.Imported_Data_From_Excel_For_Calcs[f"{self.Imported_Column_Names}"] = [value[0] for value in self.Imported_Data.values]

                self.Entry_Widget_For_W_Table_Frecuency.config(state="disabled")
                self.Value_For_Entry_Widget_W_Table_Frecuency.set(f"Columna Importada: {self.Imported_Column_Names}")

                self.Load_Selected_Data_In_Table()

    def Manage_Load_For_Module_Venn_Diagram(self):
        for data_widget in self.Entry_Widgets_For_Venn_Diagram.values():
            if(data_widget.get()):
                data_widget.set("")

        if(self.Imported_Data_From_Excel_For_Calcs):
            self.Imported_Data_From_Excel_For_Calcs.clear()

        for i , (data_widget , widget) in enumerate(zip(self.Values_For_Entry_Widgets_Venn_Diagram.values() , self.Entry_Widgets_For_Venn_Diagram.values())):
            if(i < len(self.Imported_Column_Names)):
                self.Imported_Data_From_Excel_For_Calcs[self.Imported_Column_Names[i]] = [value for value in self.Imported_Data[self.Imported_Column_Names[i]].dropna()]

                widget.config(state="disabled")
                data_widget.set(f"Columna Importada: {self.Imported_Column_Names[i]}")
            else:
                widget.config(state="disabled")

        self.Load_Selected_Data_In_Table(self.Start_Row , self.End_Row)

class Importer_Of_All_Data_In_Excel_File:
    def __init__(self , W_Import_Excel , File_Path , Sheet_Number_Intvar_Value , Table_For_Show_Imported_Data):
        """
            Esta clase permite importar todos los datos de una hoja especifica de un archivo excel.
        """
        self.W_Import_Excel = W_Import_Excel
        self.File_Path = File_Path
        self.Sheet_Number_Intvar_Value = Sheet_Number_Intvar_Value
        self.Sheet_Number_Int_Value = Sheet_Number_Intvar_Value.get()

        self.Table_For_Show_Imported_Data = Table_For_Show_Imported_Data

        self.Total_Rows_In_Excel_Sheet = 0
        self.Total_Columns_In_Excel_Sheet = 0

    def Get_Excel_File(self):
        try:
            if(not self.File_Path):
                raise Raise_Warning("No se encontro la ruta del archivo")
        
            prev_load_excel = openpyxl.load_workbook(self.File_Path , read_only=True , data_only=True , keep_links=False)
            Sheets = prev_load_excel.sheetnames

            if(self.Sheet_Number_Int_Value > len(Sheets)):
                self.W_Import_Excel.after(550 , lambda: self.Sheet_Number_Intvar_Value.set(self.Sheet_Number_Int_Value - 1))
                raise Raise_Warning(f"El numero de hoja {self.Sheet_Number_Int_Value} no existe")
            
            Idx_Sheet = self.Sheet_Number_Int_Value - 1

            Sheet_Name = Sheets[Idx_Sheet]
            One_Sheet = prev_load_excel[Sheet_Name]

            self.Total_Rows_In_Excel_Sheet = One_Sheet.max_row
            self.Total_Columns_In_Excel_Sheet = One_Sheet.max_column

            self.Get_Data_From_One_Sheet(Idx_Sheet)

        except Raise_Warning as e:
            self.W_Import_Excel.after(0 , messagebox.showwarning("Advertencia" , e))
            self.W_Import_Excel.after(30 , Insert_Data_In_Log_File(e , "Advertencia" , "Thread de importacion de datos"))
            return
        except Exception as e:
            self.W_Import_Excel.after(0 , messagebox.showerror("Error" , "Ocurrio un error al intentar importar los datos del archivo"))
            self.W_Import_Excel.after(30 , Insert_Data_In_Log_File("Ocurrio un error al intentar importar los datos del archivo" , "Error" , "Thread de importacion de datos" , Get_Detailed_Info_About_Error()))
            return
        else:
            self.W_Import_Excel.after(0 , Insert_Data_In_Log_File("Todos los datos del excel se importaron correctamente" , "Operacion exitosa" , "Thread de importacion de datos de un excel"))
            
    def Get_Data_From_One_Sheet(self , Idx_Sheet):
        JSON_Settings_Data = Read_Data_From_JSON("import_excel_settings")

        Data_Excel = CalamineWorkbook.from_path(self.File_Path).get_sheet_by_index(Idx_Sheet).to_python(skip_empty_area=False)
        if(Idx_Sheet in self.Table_For_Show_Imported_Data.Extra_Data_For_Been_Saved):
            self.Excel_Dataframe = self.Table_For_Show_Imported_Data.Extra_Data_For_Been_Saved[Idx_Sheet]["All_Excel_Sheet_Data"]
        elif(Data_Excel):
            self.Excel_Dataframe = pd.DataFrame(data=Data_Excel[1:] , columns=Data_Excel[0])
        else:
            self.Excel_Dataframe = pd.DataFrame(data=[""] , columns=["No hay datos"])

        self.First_One_Hundred_Data = self.Excel_Dataframe.head(JSON_Settings_Data["maximun_rows_to_display_in_preview"] - 1)
        self.List_Number_Data_In_Row = []
        Values_Recognized_as_Null = ["" , "NaN" , "None"]

        for col_pos in range(self.Excel_Dataframe.shape[1]):
            val = self.Excel_Dataframe.iloc[: , col_pos].to_list()
            counter_data_in_row = 0
            counter_void_places = 0
            for i , data in enumerate(val):
                if(not str(data).replace(" ","") in Values_Recognized_as_Null):
                    if(counter_void_places != 0):
                        counter_data_in_row += counter_void_places
                        counter_void_places = 0
                    
                    counter_data_in_row += 1
                else:
                    counter_void_places += 1
                
                if(counter_void_places > JSON_Settings_Data["void_tolerance"]):
                    break

            self.List_Number_Data_In_Row.append(counter_data_in_row + 1)
        
        if(not Idx_Sheet in self.Table_For_Show_Imported_Data.Extra_Data_For_Been_Saved):
            self.Table_For_Show_Imported_Data.Extra_Data_For_Been_Saved[Idx_Sheet] = {
                "All_Excel_Sheet_Data": self.Excel_Dataframe,
                "Total_Rows_In_Excel_Sheet": self.Total_Rows_In_Excel_Sheet,
                "Total_Columns_In_Excel_Sheet": self.Total_Columns_In_Excel_Sheet,
            }
        
        if(not self.File_Path in self.Table_For_Show_Imported_Data.Extra_Data_For_Been_Saved):
            self.Table_For_Show_Imported_Data.Extra_Data_For_Been_Saved["File_Path"] = self.File_Path

    def Load_Excel_In_Table(self):
        List_With_All_Columns_Letters = []
        for i in range(len(self.First_One_Hundred_Data.columns)):
            Col_Letter = index_to_string(i)
            List_With_All_Columns_Letters.append(Col_Letter)
        
        Titles_For_Columns = ["N° fila / columna"] + List_With_All_Columns_Letters
        self.Table_For_Show_Imported_Data.Modify_Number_Of_Columns(len(Titles_For_Columns) , Titles_For_Columns , 120)

        Void_Space_In_Bottom_Preview = [""] + ["" for _ in range(len(Titles_For_Columns))]

        Data_To_Display_In_Table = []
        Data_To_Display_In_Table.append(tuple([1] + self.First_One_Hundred_Data.columns.tolist()))
        for (index, data_in_row) in self.First_One_Hundred_Data.iterrows():
            Data_To_Display_In_Table.append(tuple([index + 2] + data_in_row.tolist()))

        Data_To_Display_In_Table.append(Void_Space_In_Bottom_Preview)
        
        Extra_Data_In_Bottom_Of_Table = ["Ultimo dato en:"] + [f"{col_letter}{row_count}" for col_letter , row_count in zip(List_With_All_Columns_Letters , self.List_Number_Data_In_Row)]

        self.Table_For_Show_Imported_Data.Insert_Data(Data_To_Display_In_Table , Extra_Data_In_Bottom_Of_Table)

class Selecter_Of_Data_For_Single_Range_Of_Cells(Validator , Loader_Of_Selected_Data_In_Table):
    def __init__(self , W_Import_Excel , Table_For_Show_Selected_Data , Range_Cells , Source_Module_Name , Entry_Widget , Value_For_Entry_Widget , Imported_Data_From_Excel):
        """ 
            Esta clase permite seleccionar solo los datos indicados de una hoja de excel mediante un rango
            de celdas, siempre y cuando el rango de celdas a importar sea continua y contenga mas de un dato.
            Por ejemplo, los siguientes rangos de celdas son validos:
            A1:A1001
            D1:E10000
            Pero estos rangoes serian invalidos:
            A1:A1001;C1:C1001
        """
        Validator.__init__(self)
        Loader_Of_Selected_Data_In_Table.__init__(self)

        self.Str_Range_Of_Cells = Range_Cells
        self.Import_Multiple_Columns = False

        self.Table_For_Show_Selected_Data = Table_For_Show_Selected_Data
        self.W_Import_Excel = W_Import_Excel

        self.Source_Module_Name = Source_Module_Name

        if(Source_Module_Name == "Table_Of_Frecuency"):
            self.Entry_Widget_For_W_Table_Frecuency = Entry_Widget
            self.Value_For_Entry_Widget_W_Table_Frecuency = Value_For_Entry_Widget
        elif(Source_Module_Name == "Venn_Diagram"):
            self.Entry_Widgets_For_Venn_Diagram = Entry_Widget
            self.Values_For_Entry_Widgets_Venn_Diagram = Value_For_Entry_Widget

        self.Imported_Data_From_Excel_For_Calcs = Imported_Data_From_Excel

    def Process_Input_Data(self):
        """  
            Permite Validar y Procesar los rangos de celdas ingresados.
        """

        self.Start_Column , self.Start_Row , self.End_Column , self.End_Row = Validator.Validate_Format_For_Each_Range_Cells(self , self.Str_Range_Of_Cells)

        self.Start_Row , self.End_Row = int(self.Start_Row) , int(self.End_Row)

        self.Start_Row , self.End_Row , self.Start_Column , self.End_Column = Validator.Validate_Order_For_Each_Range_Cells(self, self.Start_Row , self.End_Row , self.Start_Column , self.End_Column)

        if(self.End_Column != self.Start_Column):
            self.Import_Multiple_Columns = True

    def Select_Data_From_Excel_Dataframe(self , Loaded_Excel_Dataframe , Total_Rows_In_Excel_Sheet , Total_Columns_In_Excel_Sheet):
        try:
            Validator.Validate_Row_Limit_Excel(self, self.Start_Row , self.End_Row , Total_Rows_In_Excel_Sheet)
            Validator.Validate_Column_Limit_Excel(self , self.Start_Column , self.End_Column , Total_Columns_In_Excel_Sheet)

            if(self.Start_Column != self.End_Column):
                Idx_Columns = [idx for idx in range(string_to_index(self.Start_Column) , string_to_index(self.End_Column) + 1)]
                self.Imported_Column_Names = [col_name for i , col_name in enumerate(Loaded_Excel_Dataframe.columns) if i in Idx_Columns]
                if("" in self.Imported_Column_Names):
                    raise Raise_Warning("Se intento importar datos sin un encabezado adecuado. Por favor, coloque un nombre adecuado a los datos y coloquelos en la primera fila.")

            else:
                Idx_Columns = [string_to_index(self.Start_Column)]
                self.Imported_Column_Names = Loaded_Excel_Dataframe.columns[string_to_index(self.Start_Column)]
                if(self.Imported_Column_Names == ""):
                    raise Raise_Warning("Se intento importar datos sin un encabezado adecuado. Por favor, coloque un nombre adecuado a los datos y coloquelos en la primera fila.")

            if(self.Start_Row == 1):
                self.Imported_Data = Loaded_Excel_Dataframe.iloc[self.Start_Row-1:self.End_Row-1 , Idx_Columns]
            else:
                self.Imported_Data = Loaded_Excel_Dataframe.iloc[self.Start_Row-2:self.End_Row-1 , Idx_Columns]

            self.Imported_Data = self.Imported_Data.dropna()

            Validator.Validate_Data_Imported_Is_Null(self , self.Imported_Data)
        except RuntimeError as e:
            self.W_Import_Excel.after(0 , Insert_Data_In_Log_File("Error al procesar en hilos. Error en tiempo de ejecucion. Si ocurre demasiadas veces reportelo" , "Error" , "Thread de importacion de datos de un excel" , Get_Detailed_Info_About_Error()))
            self.W_Import_Excel.after(10 , messagebox.showerror("Error" , "Error al procesar en hilos.\nError en tiempo de ejecucion.\nSi ocurre demasiadas veces reportelo"))
            return
        except Raise_Warning as e:
            self.W_Import_Excel.after(0 , Insert_Data_In_Log_File(e , "Advertencia" , "Thread de importacion de datos de un excel"))
            self.W_Import_Excel.after(10 , messagebox.showwarning("Advertencia" , e))
            return
        except Exception as e:
            self.W_Import_Excel.after(0 , Insert_Data_In_Log_File("Ocurrio un error al intentar importar los datos del archivo" , "Error" , "Thread de importacion de datos de un excel" , Get_Detailed_Info_About_Error()))
            self.W_Import_Excel.after(10 , messagebox.showerror("Error" , "Ocurrio un error al intentar importar los datos del archivo"))
            return
        else:
            self.W_Import_Excel.after(0 , Insert_Data_In_Log_File("Se seleccionaron correctamente los datos solicitados" , "Operacion exitosa" , "Thread de importacion de datos de un excel"))
            
    def Load_Excel_Data_In_Table(self):
        try:
            match(self.Source_Module_Name):
                case "Table_Of_Frecuency":
                    self.Manage_Load_For_Module_Table_Of_Frecuency()
                case "Venn_Diagram":
                    if(len(self.Imported_Column_Names) < 2):
                        raise Raise_Warning("No se puede importar menos de 2 columnas.")
                    
                    self.Manage_Load_For_Module_Venn_Diagram()
        except Raise_Warning as e:
            Insert_Data_In_Log_File(e , "Advertencia" , "Importacion de datos de un excel")
            messagebox.showwarning("Advertencia" , e)
        except Exception:
            Insert_Data_In_Log_File("Ocurrio un error al intentar mostrar los datos del excel importado en la tabla de previsualizacion" , "Error" , "Importacion de datos de un excel" , Get_Detailed_Info_About_Error())
            messagebox.showerror("Error" , "Ocurrio un error al intentar mostrar los datos del excel importado en la tabla de previsualizacion")
        else:
            Insert_Data_In_Log_File("Los datos se insertaron correctamente a la tabla de previsualizacion" , "Operacion exitosa" , "Importacion de datos de un excel")
            messagebox.showinfo("Success" , "Datos procesados con exito.\nYa puede salir de la ventana de importacion.")

class Selecter_Of_Data_For_Multiple_Range_Of_Cells(Validator , Loader_Of_Selected_Data_In_Table):
    def __init__(self , W_Import_Excel , Table_For_Show_Selected_Data , Range_Cells , Source_Module_Name , Entry_Widget , Value_For_Entry_Widget , Imported_Data_From_Excel):
        Validator.__init__(self)
        Loader_Of_Selected_Data_In_Table.__init__(self)

        self.Str_Range_Of_Cells = Range_Cells

        self.Error_In_Thread = False
        self.Info_About_Error = []

        self.Collection_Of_Cells = {
            "Columns" : [],
            "Rows" : [],
        }

        self.Import_Multiple_Columns = True

        self.W_Import_Excel = W_Import_Excel
        self.Table_For_Show_Selected_Data = Table_For_Show_Selected_Data
        
        self.Source_Module_Name = Source_Module_Name

        if(Source_Module_Name == "Table_Of_Frecuency"):
            self.Entry_Widget_For_W_Table_Frecuency = Entry_Widget
            self.Value_For_Entry_Widget_W_Table_Frecuency = Value_For_Entry_Widget
        elif(Source_Module_Name == "Venn_Diagram"):
            self.Entry_Widgets_For_Venn_Diagram = Entry_Widget
            self.Values_For_Entry_Widgets_Venn_Diagram = Value_For_Entry_Widget

        self.Imported_Data_From_Excel_For_Calcs = Imported_Data_From_Excel

    def Process_Input_Data(self):
        Fixed_Ranges_Of_Cells = []

        Ranges_Of_Cells = self.Str_Range_Of_Cells.split(";")
        Ranges_Of_Cells = Validator.Validate_Order_For_All_Range_Cells(self, Ranges_Of_Cells)
        for ran in Ranges_Of_Cells:
            self.Start_Column , self.Start_Row , self.End_Column , self.End_Row = Validator.Validate_Format_For_Each_Range_Cells(self , ran)
            self.Start_Row , self.End_Row = int(self.Start_Row) , int(self.End_Row)

            self.Start_Row , self.End_Row , self.Start_Column , self.End_Column = Validator.Validate_Order_For_Each_Range_Cells(self , self.Start_Row , self.End_Row , self.Start_Column , self.End_Column)

            Fixed_Ranges_Of_Cells.append(f"{self.Start_Column}{self.Start_Row}:{self.End_Column}{self.End_Row}")
            self.Collection_Of_Cells["Columns"].append([self.Start_Column , self.End_Column])
            self.Collection_Of_Cells["Rows"].append([self.Start_Row , self.End_Row])
        Validator.Validate_For_Avoid_Repeated_Range_Cells(self , Fixed_Ranges_Of_Cells)
        
    def Select_Data_From_Excel_Dataframe(self , Excel_Dataframe , Total_Rows_In_Excel_Sheet , Total_Columns_In_Excel_Sheet):
        try:
            Arr_Columns = []
            Arr_Rows = []
            for row in self.Collection_Of_Cells["Rows"]:
                Validator.Validate_Row_Limit_Excel(self , row[0] , row[1] , Total_Rows_In_Excel_Sheet)
                Arr_Rows.append(row[0])
                Arr_Rows.append(row[1])
            
            for col in self.Collection_Of_Cells["Columns"]:
                Validator.Validate_Column_Limit_Excel(self , col[0] , col[1] , Total_Columns_In_Excel_Sheet)
                Arr_Columns.append(col[0])
                if(string_to_index(col[1]) - string_to_index(col[0]) > 1):
                    for c in range(string_to_index(col[0]) + 1 , string_to_index(col[1])):
                        Arr_Columns.append(index_to_string(c))
                Arr_Columns.append(col[1])

            Arr_Idx_Columns = [string_to_index(idx) for idx in Arr_Columns]
            Only_Unique_Idx_Columns = list(set(Arr_Idx_Columns))
            Only_Unique_Idx_Columns.sort()

            if(Only_Unique_Idx_Columns == 1):
                raise Raise_Warning("El rango de celdas importado es incorrecto")

            self.Imported_Column_Names = [col for i , col in enumerate(Excel_Dataframe.columns) if i in Only_Unique_Idx_Columns]
            if("" in self.Imported_Column_Names):
                raise Raise_Warning("Se intento importar datos sin un encabezado adecuado. Por favor, coloque un nombre adecuado a los datos y coloquelos en la primera fila.")

            Concat_Columns = []
            try:
                for i in range(0 , len(self.Collection_Of_Cells["Rows"])):
                    """ 
                        ********************************************************************
                        Las condicionales de aqui son importantes para evitar que se
                        importen columnas diferentes a las especificadas, ya que aqui no
                        dependo del valor de indice de de letra de cada columna, debido a
                        que cuando cargo el excel limito solamente a las columnas que me
                        interesa importar y esto genera que cada columna se deba de acceder
                        de manera secuencial
                        Por ejemplo: si ingreso A1:B1001;D1:D1001
                        Solo cargo las columnas A , B y C. 
                        Y para extraer los datos de cada columna debo dirigirme a ellos como
                        0 , 1 y 2 respectivamente.
                        ********************************************************************
                    """
                    if(string_to_index(self.Collection_Of_Cells["Columns"][i][1]) - string_to_index(self.Collection_Of_Cells["Columns"][i][0]) >= 1):
                        col = [n_col for n_col in range(string_to_index(self.Collection_Of_Cells["Columns"][i][0]) , string_to_index(self.Collection_Of_Cells["Columns"][i][1]) + 1)]
                    else:
                        col = string_to_index(self.Collection_Of_Cells["Columns"][i][0])
                    
                    if(self.Collection_Of_Cells["Rows"][i][0] == 1):
                        Adjustment_Value = 1
                    else:
                        Adjustment_Value = 2
                    
                    if(isinstance(col , list)):
                        for c in col:
                            column_i = Excel_Dataframe.iloc[self.Collection_Of_Cells["Rows"][i][0]-Adjustment_Value:self.Collection_Of_Cells["Rows"][i][1]-1 , c]
                            column_i.dropna()
                            Validator.Validate_Data_Imported_Is_Null(self , column_i , False)
                            Concat_Columns.append(column_i)
                    else:
                        column_i = Excel_Dataframe.iloc[self.Collection_Of_Cells["Rows"][i][0]-Adjustment_Value:self.Collection_Of_Cells["Rows"][i][1]-1 , col]
                        column_i.dropna()
                        Validator.Validate_Data_Imported_Is_Null(self , column_i , False)
                        Concat_Columns.append(column_i)
                self.Imported_Data = pd.concat(Concat_Columns , axis=1 , ignore_index=True , join="outer")
                self.Imported_Data.columns = self.Imported_Column_Names
            except Exception as e:
                raise Raise_Warning("Algo salio mal, asegurese de que el rango de celdas ingresado no tenga intersecciones.\nCorrecto: A1:D1001;F1:H1001 \nIncorrecto: A1:D1001;C1:E1001")
            self.Imported_Data.dropna()

            self.Start_Row = min(Arr_Rows)
            self.End_Row = max(Arr_Rows)
        except RuntimeError:
            self.W_Import_Excel.after(0 , Insert_Data_In_Log_File("Error al procesar en hilos.\nError en tiempo de ejecucion.\nSi ocurre demasiadas veces reportelo" , "Error" , "Thread de importacion de datos de un excel" , Get_Detailed_Info_About_Error()))
            self.W_Import_Excel.after(10 , messagebox.showerror("Error" , "Error al procesar en hilos. Error en tiempo de ejecucion. Si ocurre demasiadas veces reportelo"))
            return
        except Raise_Warning as e:
            self.W_Import_Excel.after(0 , Insert_Data_In_Log_File(e , "Advertencia" , "Thread de importacion de datos de un excel"))
            self.W_Import_Excel.after(10 , messagebox.showwarning("Advertencia" , e))
            return
        except Exception as e:
            self.W_Import_Excel.after(0 , Insert_Data_In_Log_File("Ocurrio un error al intentar importar los datos del archivo" , "Error" , "Thread de importacion de datos de un excel" , Get_Detailed_Info_About_Error()))
            self.W_Import_Excel.after(10 , messagebox.showerror("Error" , "Ocurrio un error al intentar importar los datos del archivo"))
            return
        else:
            self.W_Import_Excel.after(0 , Insert_Data_In_Log_File("Se seleccionaron correctamente los datos solicitados" , "Operacion exitosa" , "Thread de importacion de datos de un excel"))
        
    def Load_Excel_Data_In_Table(self):
        try:
            match(self.Source_Module_Name):
                case "Table_Of_Frecuency":
                    self.Manage_Load_For_Module_Table_Of_Frecuency()
                case "Venn_Diagram":
                    if(len(self.Imported_Column_Names) < 2):
                        raise Raise_Warning("No se puede importar menos de 2 columnas.")
                    elif(len(self.Imported_Column_Names) > 6):
                        raise Raise_Warning("No se puede importar mas de 6 columnas.")
                    
                    self.Manage_Load_For_Module_Venn_Diagram()
        except Raise_Warning as e:
            Insert_Data_In_Log_File(e , "Advertencia" , "Importacion de datos de un excel")
            messagebox.showwarning("Advertencia" , e)
        except Exception:
            Insert_Data_In_Log_File("Ocurrio un error al intentar mostrar los datos del excel importado en la tabla de previsualizacion" , "Error" , "Importacion de datos de un excel" , Get_Detailed_Info_About_Error())
            messagebox.showerror("Error" , "Ocurrio un error al intentar mostrar los datos del excel importado en la tabla de previsualizacion")
        else:
            Insert_Data_In_Log_File("Los datos se insertaron correctamente a la tabla de previsualizacion" , "Operacion exitosa" , "Importacion de datos de un excel")
            messagebox.showinfo("Success" , "Datos procesados con exito.\nYa puede salir de la ventana de importacion.")