from tkinter import messagebox
from datetime import datetime
import copy
import os
from openpyxl.styles import Alignment, Font , PatternFill , Border , Side
import pandas as pd
import numpy

def Change_Key(dictionary, old_key, new_key):
    """ No modifica el diccionarrio, sino que genera uno nuevo , pero con las claves moficiadas """
    return {clave if clave != old_key else new_key: valor for clave, valor in dictionary.items()}


class Export_Data:
    def __init__(self , Data_To_Import , Type_Of_Variable , Route , File_Name):
        self.Data_To_Import = Data_To_Import
        self.Type_Of_Variable = Type_Of_Variable
        self.Route = Route
        self.File_Name = File_Name

        self.Border_With_Double_Line_Bottom = Border(
            left=Side(border_style="thin" , color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="double", color="000000")
        )
        self.Border_With_All_Borders = Border(
            left=Side(border_style="thin" , color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        self.Fill_Header_Style = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        self.Font_Courier_New_Bold = Font(bold=True , name="Courier New" , size=11)
        self.Font_Courier_New_Red_Bold = Font(bold=True , name="Courier New" , size=11 , color="FF0000")
        self.Font_Courier_New_No_Bold = Font(bold=False , name="Courier New" , size=11)

        self.Alignment_To_Cells = Alignment(horizontal="center" , vertical="center")

    def Create_Full_Route(self , Multiple_Columns=False):
        time = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        if(self.File_Name == ""):
            match(Multiple_Columns):
                case True:
                    self.File_Name = f"Tables_Frecuences_{time}.xlsx"
                case False:
                    self.File_Name = f"Frecuences_{time}.xlsx"
        elif(not self.File_Name.lower().endswith('.xlsx')):
            match(Multiple_Columns):
                case True:
                    self.File_Name += f"_MT_{time}.xlsx"
                case False:
                    self.File_Name += f'_ST_{time}.xlsx'
        
        if (self.Route == ""):
            raise Exception("No se ha ingresado ninguna ruta de exportacion.")
        if not self.Route.endswith("/"):
            self.Route += "/"

        if(not os.path.exists(self.Route) or not os.path.isdir(self.Route)):
            raise Exception("Ruta de exportacion no valida")
        
        Full_Route = self.Route + self.File_Name
        return Full_Route

    def Create_Row_Total(self , Worksheet , Col_Total_Text , Col_fi , Col_hi , Col_hi_percent):
        new_row = Worksheet.max_row + 1

        Worksheet[f"{Col_Total_Text}{new_row}"] = "Total:"
        Worksheet[f"{Col_Total_Text}{new_row}"].font = self.Font_Courier_New_Bold
        Worksheet[f"{Col_Total_Text}{new_row}"].alignment = self.Alignment_To_Cells
        Worksheet[f"{Col_Total_Text}{new_row}"].border = self.Border_With_All_Borders

        Worksheet[f"{Col_fi[0]}{new_row}"] = f"{numpy.sum(Col_fi[1])}"
        Worksheet[f"{Col_fi[0]}{new_row}"].font = self.Font_Courier_New_Red_Bold
        Worksheet[f"{Col_fi[0]}{new_row}"].alignment = self.Alignment_To_Cells
        Worksheet[f"{Col_fi[0]}{new_row}"].border = self.Border_With_All_Borders

        Worksheet[f"{Col_hi[0]}{new_row}"] = f"{round(numpy.sum(Col_hi[1]))}"
        Worksheet[f"{Col_hi[0]}{new_row}"].font = self.Font_Courier_New_Red_Bold
        Worksheet[f"{Col_hi[0]}{new_row}"].alignment = self.Alignment_To_Cells
        Worksheet[f"{Col_hi[0]}{new_row}"].border = self.Border_With_All_Borders

        Worksheet[f"{Col_hi_percent[0]}{new_row}"] = f"{round(numpy.sum(Col_hi_percent[1]))}"
        Worksheet[f"{Col_hi_percent[0]}{new_row}"].font = self.Font_Courier_New_Red_Bold
        Worksheet[f"{Col_hi_percent[0]}{new_row}"].alignment = self.Alignment_To_Cells
        Worksheet[f"{Col_hi_percent[0]}{new_row}"].border = self.Border_With_All_Borders

    def Add_Summary_Measures(self , WorkSheet, M_Central_Tendency_And_Dispersion , M_Coefficient_Asymmetry , Start_Row , Table_Length):
        if(not (M_Central_Tendency_And_Dispersion or M_Coefficient_Asymmetry)):
            raise Exception("Error al exportar las medidas de resumen.")
        Table_Length -= 1 # Como vamos a trabajar con listas quito 1

        Arr_Col = ["O" , "P" , "Q" , "R" , "S" , "T" , "U" , "V" , "W" , "X" , "Y" , "AA" , "AB" , "AC" , "AD" , "AE" , "AF" , "AG" , "AH" , "AI" , "AJ" , "AK" , "AL" , "AM"]
        Move_Col = 0
        Move_Row = 0
        WorkSheet[f"N{Start_Row}"] = "Medidas de Tendencia Central y de Dispersion"
        WorkSheet[f"N{Start_Row}"].font = self.Font_Courier_New_Bold
        WorkSheet[f"N{Start_Row}"].alignment = self.Alignment_To_Cells

        for N_Cental_T , (key_Central_T , value_Central_T) in enumerate(M_Central_Tendency_And_Dispersion.items()):
            if(Move_Row > Table_Length and Table_Length <= 6):
                Move_Col += 3
                Move_Row = 0
        
            if(N_Cental_T != 3):
                WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"] = f"{key_Central_T}"
                WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].font = self.Font_Courier_New_Bold
                WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].alignment = self.Alignment_To_Cells
                WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].border = self.Border_With_All_Borders
                WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].fill = self.Fill_Header_Style

                WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"] = value_Central_T
                WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].font = self.Font_Courier_New_No_Bold
                WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].alignment = self.Alignment_To_Cells
                WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].border = self.Border_With_All_Borders
            else:
                if(Move_Row + len(value_Central_T) > Table_Length and len(value_Central_T) > 1):
                    Move_Col += 3
                    Move_Row = 0
                
                WorkSheet.merge_cells(f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}:{Arr_Col[Move_Col]}{Start_Row + len(value_Central_T) - 1 + Move_Row}")
                WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"] = f"{key_Central_T}"
                WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].font = self.Font_Courier_New_Bold
                WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].alignment = self.Alignment_To_Cells
                WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].border = self.Border_With_All_Borders
                WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].fill = self.Fill_Header_Style
                if(len(value_Central_T) > 1):
                    for n , Mo in enumerate(value_Central_T , start=1):
                        WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"] = f"Mo_{n} : {Mo}"
                        WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].font = self.Font_Courier_New_No_Bold
                        WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].alignment = self.Alignment_To_Cells
                        WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].border = self.Border_With_All_Borders
                        Move_Row += 1
                    Move_Row -= 1
                else:
                    WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"] = f"Mo_1 : {value_Central_T[0]}"
                    WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].font = self.Font_Courier_New_No_Bold
                    WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].alignment = self.Alignment_To_Cells
                    WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].border = self.Border_With_All_Borders
            Move_Row += 1

        
        WorkSheet[f"{Arr_Col[Move_Col + 3]}{Start_Row}"] = "Coeficientes de Asimetria"
        WorkSheet[f"{Arr_Col[Move_Col + 3]}{Start_Row}"].font = self.Font_Courier_New_Bold
        WorkSheet[f"{Arr_Col[Move_Col + 3]}{Start_Row}"].alignment = self.Alignment_To_Cells

        Arr_Col = [col for i , col in enumerate(Arr_Col) if i > Move_Col + 3]
        Move_Row = 0
        Move_Col = 0
        for key_Coef_Asymmetry , value_Coef_Asymmetry in M_Coefficient_Asymmetry.items():
            WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"] = f"Coeficiente de {key_Coef_Asymmetry}"
            WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].font = self.Font_Courier_New_Bold
            WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].alignment = self.Alignment_To_Cells
            WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].border = self.Border_With_All_Borders
            WorkSheet[f"{Arr_Col[Move_Col]}{Start_Row + Move_Row}"].fill = self.Fill_Header_Style

            WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"] = value_Coef_Asymmetry
            WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].font = self.Font_Courier_New_No_Bold
            WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].alignment = self.Alignment_To_Cells
            WorkSheet[f"{Arr_Col[Move_Col + 1]}{Start_Row + Move_Row}"].border = self.Border_With_All_Borders
            Move_Row += 1

    def Adjust_Width(self , Worksheet):
        Font_Factor = 1.3 # Para Courier New tamaño 12
        for col in Worksheet.columns:
            Max_Length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if(cell.value):
                        Cell_Length = len(str(cell.value))
                        if Cell_Length > Max_Length:
                            Max_Length = Cell_Length
                except:
                    pass
            Adjusted_Width = (Max_Length + 2) * Font_Factor
            Worksheet.column_dimensions[column].width = Adjusted_Width

    def Align_And_Style_Values_In_Table_Cells(self , Worksheet , Table_Length , Type_Of_Data = "" , Start_Row = 3):
        Arr_col = ["D" , "E" , "F" , "G" , "H" , "I" , "J"]

        if(Type_Of_Data == "Cuantitative_Grouped"):
            Arr_col = ["D" , "E" , "F" , "G" , "H" , "I" , "J" , "K"]

        for col in Arr_col:
            for row in range(Start_Row , Table_Length + Start_Row):
                Worksheet[f"{col}{row}"].alignment = self.Alignment_To_Cells
                Worksheet[f"{col}{row}"].border = self.Border_With_All_Borders
                Worksheet[f"{col}{row}"].font = self.Font_Courier_New_No_Bold

            cell = Worksheet[f"{col}{Start_Row}"] # Accedo a por ejemplo A4 o D4
            
            cell.font = self.Font_Courier_New_Red_Bold
            cell.fill = self.Fill_Header_Style
            cell.border = self.Border_With_Double_Line_Bottom

            if(col == "D"):
                for row in range(Start_Row + 1 , Worksheet.max_row + 1):
                    cell = Worksheet[f"{col}{row}"]

                    if cell.value is not None and Type_Of_Data == "Cuantitative_Grouped":
                        Old_Text = str(cell.value)
                        Old_Text = Old_Text.replace("[","").replace("]","").replace(",","").replace(" "," - ").replace("np.float64(","").replace(")","")

                        if(row != Worksheet.max_row):
                            New_Text = "[ " + Old_Text + " >"
                        else:
                            New_Text = "[ " + Old_Text + " ]"
                        cell.value = New_Text
                    
                    cell.font = self.Font_Courier_New_Bold

    def Export_Excel_With_Single_Table(self):
        Full_Route = self.Create_Full_Route()
        Key_Data_To_Import = None
        Value_Data_To_Import = None
        with pd.ExcelWriter(Full_Route) as writer:
            Workboox = writer.book
            Worksheet_1 = Workboox.create_sheet("Hoja1")

            Start_Row = 4
            if(len(self.Data_To_Import) == 1 and isinstance(self.Data_To_Import , dict)):
                Key_Data_To_Import , Value_Data_To_Import = next(iter(self.Data_To_Import.items()))
                Copy_Data = copy.deepcopy(Value_Data_To_Import)
            else:
                Copy_Data = copy.deepcopy(self.Data_To_Import)

            M_Central_Tendency_And_Dispersion = None
            M_Coefficient_Asymmetry = None
            if("Summary_Measures_For_Grouped_Data" in Copy_Data):
                M_Central_Tendency_And_Dispersion = Copy_Data["Summary_Measures_For_Grouped_Data"]["Measures_Of_Central_Tendency_And_Dispersion"]
                M_Coefficient_Asymmetry = Copy_Data["Summary_Measures_For_Grouped_Data"]["Coefficient_Asymmetry"]
            elif("Summary_Measures_For_Not_Grouped_Data" in Copy_Data):
                M_Central_Tendency_And_Dispersion = Copy_Data["Summary_Measures_For_Not_Grouped_Data"]["Measures_Of_Central_Tendency_And_Dispersion"]
                M_Coefficient_Asymmetry = Copy_Data["Summary_Measures_For_Not_Grouped_Data"]["Coefficient_Asymmetry"]

            if("Frecuences_Cuant_Grouped" in Copy_Data):
                Copy_Data =  Copy_Data["Frecuences_Cuant_Grouped"]
            elif("Frecuences_Cuant_Not_Grouped" in Copy_Data):
                Copy_Data = Copy_Data["Frecuences_Cuant_Not_Grouped"]
            elif("Frecuences_Cuali" in Copy_Data):
                Copy_Data = Copy_Data["Frecuences_Cuali"]

            Copy_Data = Change_Key(Copy_Data , "hi_percent" , "hi%")
            Copy_Data = Change_Key(Copy_Data , "Hi_percent" , "Hi%")
            if("Intervals" in Copy_Data):
                Copy_Data = Change_Key(Copy_Data , "Intervals" , "[ Li - Ls >")

            Data_Excel = pd.DataFrame(Copy_Data)
            Data_Excel.to_excel(writer , sheet_name="Hoja1" , index=False , startrow=Start_Row-1 , startcol=3)

            if(Key_Data_To_Import):
                Worksheet_1[f"C{Start_Row}"] = f"Tabla para: \"{Key_Data_To_Import}\""
                Worksheet_1[f"C{Start_Row}"].font = Font(bold=True)
                Worksheet_1[f"C{Start_Row}"].alignment = Alignment(horizontal="center" , vertical="center")

            Table_Length = len(Data_Excel) + 1

            match(self.Type_Of_Variable):
                case "Cuantitative_Grouped":
                    self.Align_And_Style_Values_In_Table_Cells(Worksheet_1 , Table_Length , self.Type_Of_Variable , Start_Row)
                    self.Create_Row_Total(Worksheet_1 , "D" , ["G" , Copy_Data["fi"]] , ["I" , Copy_Data["hi"]] , ["K" , Copy_Data["hi%"]])
                    self.Add_Summary_Measures(Worksheet_1 , M_Central_Tendency_And_Dispersion , M_Coefficient_Asymmetry , Start_Row , Table_Length)
                case "Cuantitative_Not_Grouped":
                    self.Align_And_Style_Values_In_Table_Cells(Worksheet_1 , Table_Length , self.Type_Of_Variable , Start_Row)
                    self.Create_Row_Total(Worksheet_1 , "D" , ["E" , Copy_Data["fi"]] , ["G" , Copy_Data["hi"]] , ["I" , Copy_Data["hi%"]])
                    self.Add_Summary_Measures(Worksheet_1 , M_Central_Tendency_And_Dispersion , M_Coefficient_Asymmetry , Start_Row , Table_Length)
                case "Cualitative":
                    self.Align_And_Style_Values_In_Table_Cells(Worksheet_1 , Table_Length , self.Type_Of_Variable , Start_Row)
                    self.Create_Row_Total(Worksheet_1 , "D" , ["E" , Copy_Data["fi"]] , ["G" , Copy_Data["hi"]] , ["I" , Copy_Data["hi%"]])
                case _:
                    raise Exception("No se pudo identificar el tipo de variable del cual resultan la tabla a exportar")

            self.Adjust_Width(Worksheet_1)

    def Export_Excel_With_Multiple_Tables(self , Columns_To_Export):
        Full_Route = self.Create_Full_Route(True)

        Copy_Data = copy.deepcopy(self.Data_To_Import)
        
        with pd.ExcelWriter(Full_Route) as writer:
            Workbook = writer.book
            Worksheet_1 = Workbook.create_sheet("Hoja1")
            Start_Row = 4
            for key , value in Copy_Data.items():
                if(not Columns_To_Export[key].get()):
                    continue

                M_Central_Tendency_And_Dispersion = None
                M_Coefficient_Asymmetry = None
                if("Summary_Measures_For_Grouped_Data" in value):
                    M_Central_Tendency_And_Dispersion = value["Summary_Measures_For_Grouped_Data"]["Measures_Of_Central_Tendency_And_Dispersion"]
                    M_Coefficient_Asymmetry = value["Summary_Measures_For_Grouped_Data"]["Coefficient_Asymmetry"]
                    value.pop("Summary_Measures_For_Grouped_Data")
                elif("Summary_Measures_For_Not_Grouped_Data" in value):
                    M_Central_Tendency_And_Dispersion = value["Summary_Measures_For_Not_Grouped_Data"]["Measures_Of_Central_Tendency_And_Dispersion"]
                    M_Coefficient_Asymmetry = value["Summary_Measures_For_Not_Grouped_Data"]["Coefficient_Asymmetry"]
                    value.pop("Summary_Measures_For_Not_Grouped_Data")
                
                if("Frecuences_Cuant_Grouped" in value):
                    value = value["Frecuences_Cuant_Grouped"]
                elif("Frecuences_Cuant_Not_Grouped" in value):
                    value = value["Frecuences_Cuant_Not_Grouped"]
                elif("Frecuences_Cuali" in value):
                    value = value["Frecuences_Cuali"]
                
                value = Change_Key(value , "hi_percent" , "hi%")
                value = Change_Key(value , "Hi_percent" , "Hi%")

                if("Intervals" in value):
                    value = Change_Key(value , "Intervals" , "[ Li - Ls >")

                Data_Excel = pd.DataFrame(value)

                Data_Excel.to_excel(writer , sheet_name="Hoja1" , index=False , startrow=Start_Row-1 , startcol=3) #Aqui tengo que restar uno porque pandas cuenta desde el 0

                Worksheet_1[f"C{Start_Row}"] = f"Tabla para: \"{key}\""
                Worksheet_1[f"C{Start_Row}"].font = Font(bold=True)
                Worksheet_1[f"C{Start_Row}"].alignment = Alignment(horizontal="center" , vertical="center")

                """ 
                    Total de filas que ocupa la tabla en el excel, mas 1 de la fila TOTAL, ya que pandas todavia no cuenta esas filas.
                """
                Table_Length = len(Data_Excel) + 1

                match(self.Type_Of_Variable[key]):
                    case "Cuantitative_Grouped":
                        self.Align_And_Style_Values_In_Table_Cells(Worksheet_1 , Table_Length , self.Type_Of_Variable[key] , Start_Row)
                        self.Create_Row_Total(Worksheet_1 , "D" , ["G" , value["fi"]] , ["I" , value["hi"]] , ["K" , value["hi%"]])
                        self.Add_Summary_Measures(Worksheet_1 , M_Central_Tendency_And_Dispersion , M_Coefficient_Asymmetry , Start_Row , Table_Length)
                    case "Cuantitative_Not_Grouped":
                        self.Align_And_Style_Values_In_Table_Cells(Worksheet_1 , Table_Length , self.Type_Of_Variable[key] , Start_Row)
                        self.Create_Row_Total(Worksheet_1 , "D" , ["E" , value["fi"]] , ["G" , value["hi"]] , ["I" , value["hi%"]])
                        self.Add_Summary_Measures(Worksheet_1 , M_Central_Tendency_And_Dispersion , M_Coefficient_Asymmetry , Start_Row , Table_Length)
                    case "Cualitative":
                        self.Align_And_Style_Values_In_Table_Cells(Worksheet_1 , Table_Length , self.Type_Of_Variable[key] , Start_Row)
                        self.Create_Row_Total(Worksheet_1 , "D" , ["E" , value["fi"]] , ["G" , value["hi"]] , ["I" , value["hi%"]])
                    case _:
                        raise Exception("No se pudo identificar el tipo de variable del cual resultan la tabla a exportar")
                """
                    Aqui se añade al valor de Start_Row (al inicio es 4)
                    el numero de filas de la tabla ingresada previamente y se aumenta 4 para
                    una separacion de 4 filas entre cada tabla.
                """
                Start_Row += Table_Length + 4 # 4 filas de separacion entre tablas

            self.Adjust_Width(Worksheet_1)

def Export_Table_In_Excel(
        W_Export_Excel , Results_From_Single_Column , Results_From_Multiple_Columns , Type_Of_Variable_Single_Column , Type_Of_Variable_Multiple_Column, 
        Route , Columns_To_Export , File_Name = ""
    ):
    try:
        if(Results_From_Single_Column != {}):
            For_Single_Column_Data( Results_From_Single_Column , Type_Of_Variable_Single_Column , Route , File_Name)
        elif(Results_From_Multiple_Columns != {}):
            if(all(value == False for value in Columns_To_Export)):
                raise Exception("No se ha seleccionado ninguna columna a exportar.")
            For_Multiple_Column_Data(Results_From_Multiple_Columns , Type_Of_Variable_Multiple_Column , Route , File_Name , Columns_To_Export)
        else:
            raise Exception("No se encontraron los datos a exportar.")
    except Exception as e:
        messagebox.showerror("Error" , f"Hubo un error al exportar el Excel\n{e}")
    else:
        messagebox.showinfo("Sucess" , f"Excel exportado correctamente a {Route}")
        W_Export_Excel.quit()
        W_Export_Excel.destroy()

def For_Single_Column_Data(Results_From_Single_Column , Type_Of_Variable , Route , File_Name = ""):
    Export = Export_Data(Results_From_Single_Column , Type_Of_Variable , Route , File_Name)
    Export.Export_Excel_With_Single_Table()

def For_Multiple_Column_Data(Results_From_Multiple_Columns , Type_Of_Variables , Route , File_Name , Columns_To_Export):
    Exports = Export_Data(Results_From_Multiple_Columns , Type_Of_Variables , Route , File_Name)
    Exports.Export_Excel_With_Multiple_Tables(Columns_To_Export)
