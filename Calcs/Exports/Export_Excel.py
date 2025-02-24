from tkinter import messagebox
from datetime import datetime
import copy
import os
from openpyxl.styles import Alignment, Font
import pandas as pd
import numpy

class Export_Data:
    def __init__(self , Data_To_Import , Type_Of_Variable , Route , File_Name):
        self.Data_To_Import = Data_To_Import
        self.Type_Of_Variable = Type_Of_Variable
        self.Route = Route
        self.File_Name = File_Name

    def Create_Full_Route(self , Multiple_Columns=False):
        time = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        if(self.File_Name == ""):
            match(Multiple_Columns):
                case True:
                    self.File_Name = f"Tables_Frecuences_{time}.xlsx"
                case False:
                    self.File_Name = f"Frecuences_{time}.xlsx"
        elif(not self.File_Name.lower().endswith('.xlsx')):
            match(Multiple_Columns):
                case True:
                    self.File_Name += f"_MT_{time}.xlsx"
                case False:
                    self.File_Name += f'_ST_{time}.xlsx'
        
        if (self.Route == ""):
            raise Exception("No se ha ingresado ninguna ruta de exportacion.")
        if not self.Route.endswith("/"):
            self.Route += "/"

        if(not os.path.exists(self.Route) or not os.path.isdir(self.Route)):
            raise Exception("Ruta de exportacion no valida")
        
        Full_Route = self.Route + self.File_Name
        return Full_Route

    def Create_Row_Total(self , Worksheet , Col_Total_Text , Col_fi , Col_hi , Col_hi_percent):
        new_row = Worksheet.max_row + 1

        Worksheet[f"{Col_Total_Text}{new_row}"] = "Total"
        Worksheet[f"{Col_Total_Text}{new_row}"].font = Font(bold=True , )
        Worksheet[f"{Col_Total_Text}{new_row}"].alignment = Alignment(horizontal="center" , vertical="center")

        Worksheet[f"{Col_fi[0]}{new_row}"] = f"{numpy.sum(Col_fi[1])}"
        Worksheet[f"{Col_fi[0]}{new_row}"].font = Font(bold=True, )
        Worksheet[f"{Col_fi[0]}{new_row}"].alignment = Alignment(horizontal="center" , vertical="center")

        Worksheet[f"{Col_hi[0]}{new_row}"] = f"{round(numpy.sum(Col_hi[1]))}"
        Worksheet[f"{Col_hi[0]}{new_row}"].font = Font(bold=True ,)
        Worksheet[f"{Col_hi[0]}{new_row}"].alignment = Alignment(horizontal="center" , vertical="center")

        Worksheet[f"{Col_hi_percent[0]}{new_row}"] = f"{round(numpy.sum(Col_hi_percent[1]))}"
        Worksheet[f"{Col_hi_percent[0]}{new_row}"].font = Font(bold=True , )
        Worksheet[f"{Col_hi_percent[0]}{new_row}"].alignment = Alignment(horizontal="center" , vertical="center")

    def Add_Summary_Measures(self , WorkSheet, S_Measures , Start_Row):
        if(S_Measures == None):
            raise Exception("Error al exportar las medidas de resumen.")

        Arr_Col = ["N" , "O" , "P" , "Q" , "R" , "S" , "T" , "U"]
        for N_col , Key_SM in enumerate(S_Measures.keys()):
            if(N_col != 3):
                WorkSheet[f"{Arr_Col[N_col]}{Start_Row}"] = f"{Key_SM}"
                WorkSheet[f"{Arr_Col[N_col]}{Start_Row}"].font = Font(bold=True)
                WorkSheet[f"{Arr_Col[N_col]}{Start_Row}"].alignment = Alignment(horizontal="center" , vertical="center")

                WorkSheet[f"{Arr_Col[N_col]}{Start_Row + 1}"] = S_Measures[Key_SM]
                WorkSheet[f"{Arr_Col[N_col]}{Start_Row + 1}"].alignment = Alignment(horizontal="center" , vertical="center")
            else:
                WorkSheet[f"{Arr_Col[N_col]}{Start_Row}"] = f"{Key_SM}"
                WorkSheet[f"{Arr_Col[N_col]}{Start_Row}"].font = Font(bold=True)
                WorkSheet[f"{Arr_Col[N_col]}{Start_Row}"].alignment = Alignment(horizontal="center" , vertical="center")
                if(len(S_Measures[Key_SM]) > 1):
                    for n , Mo in enumerate(S_Measures[Key_SM] , start=1):
                        WorkSheet[f"{Arr_Col[N_col]}{Start_Row + n}"] = f"Mo_{n} : {Mo}"
                        WorkSheet[f"{Arr_Col[N_col]}{Start_Row + n}"].alignment = Alignment(horizontal="center" , vertical="center")
                else:
                    WorkSheet[f"{Arr_Col[N_col]}{Start_Row + 1}"] = f"Mo_1 : {S_Measures[Key_SM][0]}"
                    WorkSheet[f"{Arr_Col[N_col]}{Start_Row + 1}"].alignment = Alignment(horizontal="center" , vertical="center")
            
    def Adjust_Width(self , Worksheet):
        for col in Worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            Worksheet.column_dimensions[column].width = adjusted_width

    def Align_And_Style_Values_In_Cells(self , Worksheet , Type_Of_Data = "" , Start_Row = 3):
        Arr_col = ["D" , "E" , "F" , "G" , "H" , "I" , "J"]

        if(Type_Of_Data == "Cuantitative_Grouped"):
            Arr_col = ["D" , "E" , "F" , "G" , "H" , "I" , "J" , "K" , "L"]

        for col in Arr_col:
            for row in Worksheet[col]:
                row.alignment = Alignment(horizontal="center", vertical="center")
            if(Start_Row == 3):
                cell = Worksheet[f"{col}4"] # Accedo a por ejemplo A4 o D4
            else:
                cell = Worksheet[f"{col}{Start_Row + 1}"]
            cell.font = Font(bold=True , color="FF0000")

            if(col == "D"):
                for row in range(Start_Row + 2 , Worksheet.max_row + 1):
                    cell = Worksheet[f"{col}{row}"]

                    if cell.value is not None and Type_Of_Data == "Cuantitative_Grouped":
                        Old_Text = str(cell.value)
                        Old_Text = Old_Text.replace("[","").replace("]","").replace(",","").replace(" "," - ")

                        if(row != Worksheet.max_row):
                            New_Text = "[ " + Old_Text + " >"
                        else:
                            New_Text = "[ " + Old_Text + " ]"
                        cell.value = New_Text
                    
                    cell.font = Font(bold=True)

    def Export_Excel_With_Single_Table(self):
        Full_Route = self.Create_Full_Route()
        Key_Data_To_Import = None
        Value_Data_To_Import = None
        with pd.ExcelWriter(Full_Route) as writer:
            Workboox = writer.book
            Worksheet_1 = Workboox.create_sheet("Hoja1")

            Start_Row = 3
            if(len(self.Data_To_Import) == 1 and isinstance(self.Data_To_Import , dict)):
                Key_Data_To_Import , Value_Data_To_Import = next(iter(self.Data_To_Import.items()))
                Copy_Data = copy.deepcopy(Value_Data_To_Import)
            else:
                Copy_Data = copy.deepcopy(self.Data_To_Import)

            S_Measures = None
            if("Summary_Measures_For_Grouped_Data" in Copy_Data):
                S_Measures = Copy_Data["Summary_Measures_For_Grouped_Data"]
            elif("Summary_Measures_For_Not_Grouped_Data" in Copy_Data):
                S_Measures = Copy_Data["Summary_Measures_For_Not_Grouped_Data"]

            if("Frecuences_Cuant_For_Many_Values" in Copy_Data):
                Copy_Data =  Copy_Data["Frecuences_Cuant_For_Many_Values"]
            elif("Frecuences_Cuant_Normal_Extended" in Copy_Data):
                Copy_Data = Copy_Data["Frecuences_Cuant_Normal_Extended"]
            elif("Frecuences_Cuali_Normal_Extended" in Copy_Data):
                Copy_Data = Copy_Data["Frecuences_Cuali_Normal_Extended"]

            Copy_Data = Change_Key(Copy_Data , "hi_percent" , "hi%")
            Copy_Data = Change_Key(Copy_Data , "Hi_percent" , "Hi%")
            if("Intervals" in Copy_Data):
                Copy_Data = Change_Key(Copy_Data , "Intervals" , "[ Li - Ls >")

            Data_Excel = pd.DataFrame(Copy_Data)
            Data_Excel.to_excel(writer , sheet_name="Hoja1" , index=False , startrow=Start_Row , startcol=3)

            if(Key_Data_To_Import):
                Worksheet_1[f"C{Start_Row + 1}"] = f"Tabla para: \"{Key_Data_To_Import}\""
                Worksheet_1[f"C{Start_Row + 1}"].font = Font(bold=True)
                Worksheet_1[f"C{Start_Row + 1}"].alignment = Alignment(horizontal="center" , vertical="center")

            match(self.Type_Of_Variable):
                case "Cuantitative_Grouped":
                    self.Align_And_Style_Values_In_Cells(Worksheet_1 , self.Type_Of_Variable , Start_Row)
                    self.Create_Row_Total(Worksheet_1 , "D" , ["G" , Copy_Data["fi"]] , ["I" , Copy_Data["hi"]] , ["K" , Copy_Data["hi%"]])
                    self.Add_Summary_Measures(Worksheet_1 , S_Measures , Start_Row + 1)
                case "Cuantitative_Not_Grouped":
                    self.Align_And_Style_Values_In_Cells(Worksheet_1 , self.Type_Of_Variable , Start_Row)
                    self.Create_Row_Total(Worksheet_1 , "D" , ["E" , Copy_Data["fi"]] , ["G" , Copy_Data["hi"]] , ["I" , Copy_Data["hi%"]])
                    self.Add_Summary_Measures(Worksheet_1 , S_Measures , Start_Row + 1)
                case "Cualitative":
                    self.Align_And_Style_Values_In_Cells(Worksheet_1 , self.Type_Of_Variable , Start_Row)
                    self.Create_Row_Total(Worksheet_1 , "D" , ["E" , Copy_Data["fi"]] , ["G" , Copy_Data["hi"]] , ["I" , Copy_Data["hi%"]])
                case _:
                    raise Exception("No se pudo identificar el tipo de variable del cual resultan la tabla a exportar")

            Worksheet_1.auto_filter.ref = Worksheet_1.dimensions
            for col in Worksheet_1.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:  # Verifica si el valor de la celda no es None
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except Exception:
                        pass
                adjusted_width = (max_length + 2)
                Worksheet_1.column_dimensions[column].width = adjusted_width

    def Export_Excel_With_Multiple_Tables(self , Columns_To_Export):
        Full_Route = self.Create_Full_Route(True)

        Copy_Data = copy.deepcopy(self.Data_To_Import)
        
        with pd.ExcelWriter(Full_Route) as writer:
            Workbook = writer.book
            Worksheet_1 = Workbook.create_sheet("Hoja1")
            Start_Row = 3
            b = 1
            for key , value in Copy_Data.items():
                if(not Columns_To_Export[key].get()):
                    continue

                S_Measures = None
                if("Summary_Measures_For_Grouped_Data" in value):
                    S_Measures = value["Summary_Measures_For_Grouped_Data"]
                    value.pop("Summary_Measures_For_Grouped_Data")
                elif("Summary_Measures_For_Not_Grouped_Data" in value):
                    S_Measures = value["Summary_Measures_For_Not_Grouped_Data"]
                    value.pop("Summary_Measures_For_Not_Grouped_Data")
                
                if("Frecuences_Cuant_For_Many_Values" in value):
                    value = value["Frecuences_Cuant_For_Many_Values"]
                elif("Frecuences_Cuant_Normal_Extended" in value):
                    value = value["Frecuences_Cuant_Normal_Extended"]
                elif("Frecuences_Cuali_Normal_Extended" in value):
                    value = value["Frecuences_Cuali_Normal_Extended"]
                
                value = Change_Key(value , "hi_percent" , "hi%")
                value = Change_Key(value , "Hi_percent" , "Hi%")

                if("Intervals" in value):
                    value = Change_Key(value , "Intervals" , "[ Li - Ls >")

                Data_Excel = pd.DataFrame(value)

                Data_Excel.to_excel(writer , sheet_name="Hoja1" , index=False , startrow=Start_Row , startcol=3)

                Worksheet_1[f"C{Start_Row + 1}"] = f"Tabla para: \"{key}\""
                Worksheet_1[f"C{Start_Row + 1}"].font = Font(bold=True)
                Worksheet_1[f"C{Start_Row + 1}"].alignment = Alignment(horizontal="center" , vertical="center")

                Rows_In_Table = len(Data_Excel) + 3

                match(self.Type_Of_Variable[key]):
                    case "Cuantitative_Grouped":
                        self.Align_And_Style_Values_In_Cells(Worksheet_1 , self.Type_Of_Variable[key] , Start_Row)
                        self.Create_Row_Total(Worksheet_1 , "D" , ["G" , value["fi"]] , ["I" , value["hi"]] , ["K" , value["hi%"]])
                        self.Add_Summary_Measures(Worksheet_1 , S_Measures , Start_Row + 1)
                    case "Cuantitative_Not_Grouped":
                        self.Align_And_Style_Values_In_Cells(Worksheet_1 , self.Type_Of_Variable[key] , Start_Row)
                        self.Create_Row_Total(Worksheet_1 , "D" , ["E" , value["fi"]] , ["G" , value["hi"]] , ["I" , value["hi%"]])
                        self.Add_Summary_Measures(Worksheet_1 , S_Measures , Start_Row + 1)
                    case "Cualitative":
                        self.Align_And_Style_Values_In_Cells(Worksheet_1 , self.Type_Of_Variable[key] , Start_Row)
                        self.Create_Row_Total(Worksheet_1 , "D" , ["E" , value["fi"]] , ["G" , value["hi"]] , ["I" , value["hi%"]])
                    case _:
                        raise Exception("No se pudo identificar el tipo de variable del cual resultan la tabla a exportar")

                Start_Row += Rows_In_Table + 4

            Worksheet_1.auto_filter.ref = Worksheet_1.dimensions
            for col in Worksheet_1.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:  # Verifica si el valor de la celda no es None
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except Exception:
                        pass
                adjusted_width = (max_length + 2)
                Worksheet_1.column_dimensions[column].width = adjusted_width

def Change_Key(dictionary, old_key, new_key):
    """ No modifica el diccionarrio, sino que genera uno nuevo , pero con las claves moficiadas """
    return {clave if clave != old_key else new_key: valor for clave, valor in dictionary.items()}

def Export_Table_In_Excel(
        W_Export_Excel , Results_From_Single_Column , Results_From_Multiple_Columns , Type_Of_Variable_Single_Column , Type_Of_Variable_Multiple_Column, 
        Route , Columns_To_Export , File_Name = ""
    ):
    try:
        if(Results_From_Single_Column != {}):
            For_Single_Column_Data( Results_From_Single_Column , Type_Of_Variable_Single_Column , Route , File_Name)
        elif(Results_From_Multiple_Columns != {}):
            if(all(value == False for value in Columns_To_Export)):
                raise Exception("No se ha seleccionado ninguna columna a exportar.")
            For_Multiple_Column_Data(Results_From_Multiple_Columns , Type_Of_Variable_Multiple_Column , Route , File_Name , Columns_To_Export)
        else:
            raise Exception("No se encontraron los datos a exportar.")
    except Exception as e:
        messagebox.showerror("Error" , f"Hubo un error al exportar el Excel\n{e}")
    else:
        messagebox.showinfo("Sucess" , f"Excel exportado correctamente a {Route}")
        W_Export_Excel.quit()
        W_Export_Excel.destroy()

def For_Single_Column_Data(Results_From_Single_Column , Type_Of_Variable , Route , File_Name = ""):
    Export = Export_Data(Results_From_Single_Column , Type_Of_Variable , Route , File_Name)
    Export.Export_Excel_With_Single_Table()

def For_Multiple_Column_Data(Results_From_Multiple_Columns , Type_Of_Variables , Route , File_Name , Columns_To_Export):
    Exports = Export_Data(Results_From_Multiple_Columns , Type_Of_Variables , Route , File_Name)
    Exports.Export_Excel_With_Multiple_Tables(Columns_To_Export)
